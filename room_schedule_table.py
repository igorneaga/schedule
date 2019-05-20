import os
import re
import string

import datetime
import openpyxl
from openpyxl.comments import Comment
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.styles.borders import Border, Side
from openpyxl.worksheet.pagebreak import Break


class MasterDesign:

    def __init__(self, list_dict_courses, list_different_date, days, year, table_name, table_semester):
        self.list_dict_courses = list_dict_courses
        self.different_date_courses = list_different_date
        self.days = days
        self.table_year = year
        self.table_name = table_name
        self.table_semester = table_semester

        self.workbook = None
        self.sheet = None

        # College of Business rooms
        self.cob_rooms = ["MH 0102", "MH 0208", "MH 0209", "MH 0210", "MH 0211", "AH 0205", "AH 0209", "AH 0216",
                          "AH 0220", "AH 0320"]

        self.course_types_list = []

        self.main_class_controller()

    def main_class_controller(self):
        self.create_master_file()

        # Split course type by type
        classroom_courses_list = []
        hybrid_courses_list = []
        online_courses_list = []
        mba_macc_list = []
        telepresence_courses_list = []
        hubbard_courses_list = []
        error_courses_list = []
        for courses_len in range(len(self.list_dict_courses)):
            for course_type in self.list_dict_courses[courses_len].get("Type"):
                if course_type == "Classroom":
                    classroom_courses_list.append(self.list_dict_courses[courses_len])

                if course_type == "Hybrid":
                    hybrid_courses_list.append(self.list_dict_courses[courses_len])

                if (course_type == "Online") & (self.list_dict_courses[courses_len].get("Course") != "NONE None-None"):
                    online_courses_list.append(self.list_dict_courses[courses_len])

                if course_type == "MBA":
                    mba_macc_list.append(self.list_dict_courses[courses_len])

                elif course_type == "MACC":
                    mba_macc_list.append(self.list_dict_courses[courses_len])

                if course_type == "Telepresence":
                    telepresence_courses_list.append(self.list_dict_courses[courses_len])

                if course_type == "Hubbard":
                    hubbard_courses_list.append(self.list_dict_courses[courses_len])

                if course_type == "Error":
                    error_courses_list.append(self.list_dict_courses[courses_len])

        # Classroom table section
        self.classroom_table(classroom_courses_list, "Classroom Table", "Classroom Table", True)
        if not self.different_date_courses:
            pass
        else:
            sheet_name = "Classroom - 2nd Session"
            heading = "Classroom Table - Second Session " + self.different_date_courses[0].get("Start_Date").strftime(
                "%m-%d-%Y")
            self.classroom_table(self.different_date_courses, sheet_name, heading, False)

        # Hybrid table section
        if hybrid_courses_list:
            self.create_excel_sheet(sheet_name="Hybrid Table")
            self.set_excel_heading(heading_name="Hybrid Table")
            self.standard_table(hybrid_courses_list, "Hybrid Courses", 'cdeae6')

        # Online table section
        if online_courses_list:
            self.create_excel_sheet(sheet_name="Online Table")
            self.set_excel_heading(heading_name="Online Table")
            self.standard_table(online_courses_list, "Online Courses", "d5d1e7")

        # MBA MACC table section
        if mba_macc_list:
            self.create_excel_sheet(sheet_name="MBA MACC Table")
            self.set_excel_heading(heading_name="MBA MACC Table")
            self.standard_table(mba_macc_list, "MBA/MACC Courses", "bfa0bc")

        # Telepresence table section
        if telepresence_courses_list:
            self.create_excel_sheet(sheet_name="Telepresence Table")
            self.set_excel_heading(heading_name="Telepresence Table")
            self.standard_table(telepresence_courses_list, "Telepresence Courses", "cdeae6")

        # Hubbard table section
        if hubbard_courses_list:
            self.create_excel_sheet(sheet_name="Hubbard Table")
            self.set_excel_heading(heading_name="Hubbard Table")
            self.standard_table(hubbard_courses_list, "Hubbard Courses", "eebe95")

        # Not Included Courses table section
        if error_courses_list:
            self.create_excel_sheet(sheet_name="Not Included Courses")
            self.set_excel_heading(heading_name="Not Included Courses")
            self.standard_table(error_courses_list, "Telepresence Courses", "f6bfd6")

        self.save_excel_file()

    def standard_table(self, list_dict, course_type, color):
        """Creates a simple design table"""
        self.course_types_list = []

        self.sheet["A1"].fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
        self.sheet["A2"].fill = PatternFill(start_color=color, end_color=color, fill_type='solid')

        def insert_table_type(sheet, table_type):
            """Inserts table type on A2"""
            sheet.merge_cells("A2:B2")
            sheet["A2"] = table_type + ":"
            sheet["A2"].font = Font(sz=11, bold=True, italic=False)
            sheet["A2"].alignment = Alignment(horizontal='center', vertical='center')

        def insert_table_headings(sheet):
            """Inserts column headers on A3"""
            def insert_heading(heading_column, title):
                sheet[heading_column + "3"] = title
                sheet[heading_column + "3"].font = Font(name="Arial", sz=11, bold=True, italic=False)
                sheet[heading_column + "3"].alignment = Alignment(horizontal='left', vertical='center')

            insert_heading("A", "Course#/Sec.")
            insert_heading("B", "Cr.")
            insert_heading("C", "Title of course")
            insert_heading("D", "Days")
            insert_heading("E", "Time")
            insert_heading("F", "Room")
            insert_heading("G", "Faculty")
            insert_heading("H", "Dates")
            insert_heading("I", "Cap.")

        def insert_courses(sheet, excel_data):
            """Inserts dictionary data to excel cell"""

            for i in range(len(excel_data)):
                d = []
                row = i + 4
                sheet["A" + str(row)] = excel_data[i].get("Course")
                sheet["B" + str(row)] = excel_data[i].get("Credits")
                sheet["C" + str(row)] = excel_data[i].get("Course_Title")
                if excel_data[i].get("Course_Days"):
                    for days in excel_data[i].get("Course_Days"):
                        if days[0:2] == "Mo":
                            d.append("M")
                        elif days[0:2] == "Tu":
                            d.append("T")
                        elif days[0:2] == "We":
                            d.append("W")
                        elif days[0:2] == "Th":
                            d.append("H")
                        elif days[0:2] == "Fr":
                            d.append("F")
                        else:
                            d.append(days)
                sheet["D" + str(row)] = ''.join(d)
                sheet["E" + str(row)] = excel_data[i].get("Start_Time") + "-" + excel_data[i].get("End_Time")
                sheet["F" + str(row)] = excel_data[i].get("Room")
                sheet["G" + str(row)] = excel_data[i].get("Faculty")
                if isinstance(excel_data[i].get("Start_Date"), datetime.date) is True:
                    sheet["H" + str(row)] = excel_data[i].get("Start_Date").strftime('%m/%d/%Y') + "-" + \
                                            excel_data[i].get("End_Date").strftime('%m/%d/%Y')
                else:
                    sheet["H" + str(row)] = str(excel_data[i].get("Start_Date")) + "-" + \
                                            str(excel_data[i].get("End_Date"))
                sheet["I" + str(row)] = excel_data[i].get("Enrollment")

        for course_len in range(len(list_dict)):
            start_row = str(course_len + 4)
            self.color_cell(list_dict[course_len].get("Department"), "A" + start_row)
            # Fills the color
            for column in range(8):
                if course_len % 2 == 0:
                    get_cell_cord = ''.join(string.ascii_uppercase[column + 1]) + start_row
                    self.sheet[get_cell_cord].fill = PatternFill(start_color=color,
                                                                 end_color=color, fill_type='solid')

        insert_table_type(self.sheet, course_type)
        insert_table_headings(self.sheet)
        insert_courses(self.sheet, list_dict)
        # Excel table design
        self.border_all_cells("A3")
        self.color_cell_meaning(row_num=4)
        self.adjust_cells_width(False)
        self.set_page_break()

    def classroom_table(self, list_dict, name, heading, first):
        # Sets variable to empty
        self.course_types_list = []

        self.create_excel_sheet(sheet_name=name, first_sheet=first)
        list_unique_times = self.set_time_row(list_dict)
        self.set_courses(list_dict, list_unique_times)

        self.color_cell_meaning(row_num=2)

        self.set_excel_heading(heading_name=heading)
        self.adjust_cells_width()
        self.set_page_break()

    def save_excel_file(self):
        """Saves excel file by using user input"""
        if self.table_name[-5:] == ".xlsx":
            self.workbook.save('__excel_files\\' + self.table_name)
        else:
            self.table_name = "".join(self.table_name.split())
            if not self.table_name:
                self.table_name = "Empty_Name"
            self.workbook.save('__excel_files\\' + self.table_name + ".xlsx")

    def create_master_file(self):
        """Creates a folder and excel file"""
        def create_directory():
            """Creates directory for created excel file"""
            if not os.path.exists('__excel_files'):
                os.makedirs('__excel_files')
        create_directory()

        self.workbook = openpyxl.Workbook()

    def create_excel_sheet(self, sheet_name, first_sheet=False):
        """Creates or renames the sheet name"""
        if first_sheet is True:
            self.sheet = self.workbook.get_sheet_by_name('Sheet')
            self.sheet.title = sheet_name
        else:
            self.workbook.create_sheet(sheet_name)
            self.sheet = self.workbook.get_sheet_by_name(sheet_name)

    def set_excel_heading(self, heading_name):
        """Sets excel heading based on a user input"""

        self.sheet.oddHeader.center.text = str(heading_name) + " of " + str(self.table_semester) + " " + str(self.
                                                                                                             table_year)
        self.sheet.oddHeader.center.size = 14

        def set_course_term(semester, year, sheet):
            """Sets a course semester and year on a first cell"""
            sheet.merge_cells("A1:B1")
            sheet["A1"] = "Term: " + str(semester) + " " + str(year)
            sheet["A1"].font = Font(sz=11, bold=True, italic=False)
            sheet["A1"].alignment = Alignment(horizontal='center', vertical='center')
        set_course_term(self.table_semester, self.table_year, self.sheet)

    def set_time_row(self, list_dict):
        """Creates a row of unique times. Needs only for classroom table"""

        def prepare_row_time(d_list):
            """Prepares time to have a valid order and valid value. Needs only for classroom table"""
            list_times = []
            # Takes out of Online courses
            for i in range(len(d_list)):
                if d_list[i].get("Start_Time") != "Online":
                    list_times.append(d_list[i].get("Start_Time"))
                if d_list[i].get("End_Time") != "Online":
                    list_times.append(d_list[i].get("End_Time"))

            # Takes out duplicates
            list_times = list(set(list_times))

            def set_time_order(time_list):
                """Making time be on the correct order."""
                morning_time = []
                evening_time = []
                for time in range(len(time_list)):
                    # Earliest class starts at 8
                    if any(c in time_list[time][0:2] for c in ("08", "09", "10", "11", "12")):
                        morning_time.append(time_list[time][0:5])
                    # The latest class can start at 6 or 7
                    if any(c in time_list[time][0:2] for c in ("01", "02", "03", "04", "05", "06", "07")):
                        evening_time.append(time_list[time][0:5])

                morning_time.sort()
                evening_time.sort()
                # Combines them
                row_time = morning_time + evening_time
                return row_time

            list_times = set_time_order(list_times)
            return list_times

        unique_times = []
        list_unique_times = prepare_row_time(list_dict)
        temp_time_dict = dict()
        time_row_column = 2
        for t in range(len(list_unique_times)):
            alphabet = ''.join(string.ascii_uppercase[time_row_column])
            time_row = str(alphabet) + '1'
            self.sheet[time_row] = list_unique_times[t]
            self.sheet[time_row].font = Font(sz=11, bold=True, italic=False)
            self.sheet[time_row].alignment = Alignment(horizontal='center', vertical='center')
            temp_time_dict["Time"] = list_unique_times[t]
            temp_time_dict["Column_Num"] = time_row_column
            unique_times.append(temp_time_dict.copy())
            time_row_column += 1
        return unique_times

    def adjust_cells_width(self, classroom_table=True):
        """Adjusts all the cell width. It is really cool"""
        # Gets last column
        excel_max_column = self.sheet.max_column

        # Transfers column to an alphabetical format
        col_letter = ''.join(string.ascii_uppercase[excel_max_column - 4])

        for column in self.sheet.columns:
            max_length = 0
            # Gets column coordinates
            get_column = column[0].column
            # Column "A" and "B" will have a standard size
            if get_column is "A":
                if classroom_table is True:
                    self.sheet.column_dimensions["A"].width = 12
                else:
                    self.sheet.column_dimensions["A"].width = 18
            elif get_column is "B":
                if classroom_table is True:
                    self.sheet.column_dimensions["B"].width = 12
                else:
                    self.sheet.column_dimensions["B"].width = 6

            else:
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except TypeError:
                        pass
                # A formula for auto adjusted width
                adjusted_width = (max_length + 2) * 1.05
                if classroom_table is True:
                    # Limits adjusted width
                    if get_column is col_letter:
                        # Last column needs to be bigger to fit everything correctly
                        if adjusted_width > 24:
                            adjusted_width = 24
                    elif adjusted_width > 14:
                        adjusted_width = 14
                self.sheet.column_dimensions[get_column].width = adjusted_width

    def merge_excel_cells(self, start_row, start_column, end_row, end_column, style=False, bold=False):
        """Merges excel cells"""
        excel_sheet = self.sheet
        excel_sheet.merge_cells(start_row=start_row, start_column=start_column, end_row=end_row, end_column=end_column)

        def style_excel_cell(sheet, row, column):
            """Styles a cell"""
            sheet.cell(row=row, column=column).font = Font(sz=11, bold=bold, italic=False)
            sheet.cell(row=row, column=column).alignment = Alignment(horizontal='center',
                                                                     vertical='center', wrap_text=True)
        if style is True:
            style_excel_cell(excel_sheet, start_row, start_column)

    def border_all_cells(self, start_cell):
        """Borders all table"""
        # Gets table size
        excel_max_row = self.sheet.max_row
        excel_max_column = self.sheet.max_column

        # Style of a border
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))

        # Goes over each cell and applies border
        for column in range(excel_max_column):
            # Transfers column to an alphabetical format
            col_letter = ''.join(string.ascii_uppercase[column])
            for row in range(excel_max_row):
                row += 1
                self.sheet[col_letter + str(row)].border = thin_border

    def set_courses(self, list_dict, unique_times):
        """Sets courses in a excel file"""

        # An excel row number where courses can be placed
        start_excel_row = 2

        def set_room_dict(list_courses):
            """Creates a new dictionary where the key is a room number and contains all courses related to it"""
            result = dict()
            for course_len in range(len(list_courses)):
                for k, course_value in list_courses[course_len].items():
                    if k == "Room":
                        result.setdefault(course_value, [])
                        result[course_value].append(list_courses[course_len])
            return result

        def set_room_number(sheet, room_row, room):
            """Sets a room in a "A" column"""
            sheet['A' + str(room_row)] = room

        def set_days_row(sheet, day_row, days, b_num):
            """Sets days in a "B" column """
            sheet['B' + str(day_row)] = days[b_num]

        def get_cell_value(get_column, sheet, get_row):
            """Returns a cell value"""
            return sheet[get_column + str(get_row)].value

        # Creates a dictionary based on a room key
        room_course_dict = set_room_dict(list_dict)

        # Goes over room dict
        for key, value in room_course_dict.items():
            if not (key == "ONLINE" or key == "NONE" or key == "ARR"):
                if key in self.cob_rooms:
                    pass
                else:
                    # Marks a room if it is not a part of a college of business rooms.
                    clr = PatternFill(start_color='d5d8e0', end_color='d5d8e0', fill_type='solid')
                    self.sheet["A" + str(start_excel_row)].fill = clr
                    self.sheet["A" + str(start_excel_row)].comment = \
                        Comment("This room is not a part of a College of Business rooms", author="TableMaker")
                # Inserts rooms
                set_room_number(self.sheet, start_excel_row, key)

                days_len = len(self.days)
                for d in range(days_len):
                    # Inserts days
                    set_days_row(self.sheet, start_excel_row, self.days, d)
                    start_excel_row += 1
                for l in range(len(value)):
                    get_day = start_excel_row - days_len
                    # Merging cells with a bold
                    self.merge_excel_cells(get_day, 1, start_excel_row - 1, 1, True, bold=True)
                    while get_day != start_excel_row:
                        if any(c in get_cell_value('B', self.sheet, get_day) for c in (value[l].get('Course_Days'))):
                            for t in range(len(unique_times)):
                                column = ''.join(string.ascii_uppercase[unique_times[t].get("Column_Num")])
                                row = str(get_day)
                                if unique_times[t].get("Time") == value[l].get("Start_Time"):
                                    for en in unique_times:
                                        if en.get("Time") == value[l].get("End_Time"):
                                            value[l].setdefault("Cell", []).append(column + row + ":" + ''.join(
                                                string.ascii_uppercase[en.get("Column_Num")] + row))

                        get_day += 1

                    # Checking before merging
                    if value[l].get("Cell") is not None:
                        def split_cell_value(val, i):
                            return re.split('(\\d+)', val.get("Cell")[i])

                        def inset_cell_value(sheet, val, val_index, cell_one, index_one, index_two):
                            """Inserts a value into a cell with a comment if exist"""
                            if sheet[cell_one[index_one] + cell_one[index_two]].value is None:
                                if val[val_index].get("Time_Comment") is None:
                                    sheet[cell_one[index_one] + cell_one[index_two]] = val[val_index].get("Course")
                                else:
                                    sheet[cell_one[index_one] + cell_one[index_two]] =\
                                        val[val_index].get("Course") + val[val_index].get("Time_Comment")
                            else:
                                if val[val_index].get("Time_Comment") is None:
                                    sheet[cell_one[index_one] + cell_one[index_two]] = \
                                        sheet[cell_one[index_one] + cell_one[index_two]].value + " /" +\
                                        val[val_index].get("Course")
                                else:
                                    sheet[cell_one[index_one] + cell_one[index_two]] = \
                                        sheet[cell_one[index_one] + cell_one[index_two]].value + " /" + \
                                        val[val_index].get("Course") + "\n" + val[val_index].get("Time_Comment")

                        def merge_one_row(one_sheet, one_cell):
                            one_sheet.merge_cells(one_cell[0] + one_cell[1] + one_cell[2] + one_cell[3])
                            return True

                        def merge_two_rows(merge_two_sheet, merge_two_cell_one, merge_two_cell_two, do=True):
                            if (int(merge_two_cell_one[1]) == int(merge_two_cell_two[1]) - 1) & \
                                    (int(merge_two_cell_one[3]) == int(merge_two_cell_two[3]) - 1):
                                merge_two_sheet.merge_cells(merge_two_cell_one[0] + merge_two_cell_one[1] +
                                                            merge_two_cell_two[2] + merge_two_cell_two[3])
                                return True
                            else:
                                if do is True:
                                    if merge_one_row(merge_two_sheet, merge_two_cell_one) is True:
                                        merge_one_row(merge_two_sheet, merge_two_cell_two)
                                        return [merge_two_cell_one, merge_two_cell_two]

                        def merge_three_rows(merge_three_sheet, merge_three_cell_one, merge_three_cell_two,
                                             merge_three_cell_three):
                            # Merges all three rows
                            if (int(merge_three_cell_one[1]) == int(merge_three_cell_two[1]) - 1) & \
                                    (int(merge_three_cell_one[3]) == int(merge_three_cell_two[3]) - 1) & (
                                    int(merge_three_cell_two[1]) == int(merge_three_cell_three[1]) - 1) & (
                                    int(merge_three_cell_two[3]) == int(merge_three_cell_three[3]) - 1):
                                merge_three_sheet.merge_cells(
                                    merge_three_cell_one[0] + merge_three_cell_one[1] + merge_three_cell_three[2] +
                                    merge_three_cell_three[3])
                                return True
                            elif merge_two_rows(merge_three_sheet, merge_three_cell_one, merge_three_cell_two, False) \
                                    is True:
                                merge_one_row(merge_three_sheet, merge_three_cell_three)
                                return [merge_three_cell_one, merge_three_cell_three]
                            elif merge_two_rows(merge_three_sheet, merge_three_cell_two, merge_three_cell_three, False)\
                                    is True:
                                merge_one_row(merge_three_sheet, merge_three_cell_one)
                                return [merge_three_cell_two, merge_three_cell_one]
                            else:
                                if merge_one_row(merge_three_sheet, merge_three_cell_one) is True:
                                    merge_one_row(merge_three_sheet, merge_three_cell_two)
                                    merge_one_row(merge_three_sheet, merge_three_cell_three)
                                    return [merge_three_cell_two, merge_three_cell_one, merge_three_cell_three]

                        def merge_four_rows(merge_four_sheet, merge_four_cell_one, merge_four_cell_two,
                                            merge_four_cell_three, merge_four_cell_four):
                            # Merges four rows
                            if (int(merge_four_cell_one[1]) == int(merge_four_cell_two[1]) - 1) & \
                                    (int(merge_four_cell_one[3]) == int(merge_four_cell_two[3]) - 1) & (
                                    int(merge_four_cell_two[1]) == int(merge_four_cell_three[1]) - 1) & (
                                    int(merge_four_cell_two[3]) == int(merge_four_cell_three[3]) - 1) & (
                                    int(merge_four_cell_three[1]) == int(merge_four_cell_four[1]) - 1) & (
                                    int(merge_four_cell_three[3]) == int(merge_four_cell_four[3]) - 1):
                                merge_four_sheet.merge_cells(
                                    merge_four_cell_one[0] + merge_four_cell_one[1] + merge_four_cell_four[2] +
                                    merge_four_cell_four[3])
                                return True
                            elif (int(merge_four_cell_one[1]) == int(merge_four_cell_two[1]) - 1) & \
                                    (int(merge_four_cell_one[3]) == int(merge_four_cell_two[3]) - 1) & (
                                    int(merge_four_cell_two[1]) == int(merge_four_cell_three[1]) - 2) & (
                                    int(merge_four_cell_two[3]) == int(merge_four_cell_three[3]) - 2) & (
                                    int(merge_four_cell_three[1]) == int(merge_four_cell_four[1]) - 1) & (
                                    int(merge_four_cell_three[3]) == int(merge_four_cell_four[3]) - 1):
                                merge_two_rows(merge_four_sheet, merge_four_cell_one, merge_four_cell_two)
                                merge_two_rows(merge_four_sheet, merge_four_cell_three, merge_four_cell_four)
                                return [merge_four_cell_one, merge_four_cell_three]
                            elif merge_three_rows(merge_four_sheet, merge_four_cell_one, merge_four_cell_two,
                                                  merge_four_cell_three) is True:
                                merge_one_row(merge_four_sheet, merge_four_cell_four)
                                return [merge_four_cell_one, merge_four_cell_four]
                            elif merge_three_rows(merge_four_sheet, merge_four_cell_two, merge_four_cell_three,
                                                  merge_four_cell_four) is True:
                                merge_one_row(merge_four_sheet, merge_four_cell_one)
                                return [merge_four_cell_two, merge_four_cell_one]

                        def merge_five_rows(merge_five_sheet, merge_five_cell_one, merge_five_cell_two,
                                            merge_five_cell_three, merge_five_cell_four, merge_five_cell_five):
                            # Merges five rows. Needs to test once 5 days will be an option
                            if (int(merge_five_cell_one[1]) == int(merge_five_cell_two[1]) - 1) & \
                                    (int(merge_five_cell_one[3]) == int(merge_five_cell_two[3]) - 1) & (
                                    int(merge_five_cell_two[1]) == int(merge_five_cell_three[1]) - 1) & (
                                    int(merge_five_cell_two[3]) == int(merge_five_cell_three[3]) - 1) & (
                                    int(merge_five_cell_three[1]) == int(merge_five_cell_four[1]) - 1) & (
                                    int(merge_five_cell_three[3]) == int(merge_five_cell_four[3]) - 1) & (
                                    int(merge_five_cell_four[1]) == int(merge_five_cell_five[1]) - 1) & (
                                    int(merge_five_cell_four[3]) == int(merge_five_cell_five[3]) - 1):
                                merge_five_sheet.merge_cells(
                                    merge_five_cell_one[0] + merge_five_cell_one[1] + merge_five_cell_five[2] +
                                    merge_five_cell_five[3])
                                return True
                            elif merge_four_rows(merge_five_sheet, merge_five_cell_one, merge_five_cell_two,
                                                 merge_five_cell_three, merge_five_cell_four) is True:
                                merge_one_row(merge_five_sheet, merge_five_cell_five)
                                return [merge_five_cell_one, merge_five_cell_five]
                            elif merge_four_rows(merge_five_sheet, merge_five_cell_two, merge_five_cell_three,
                                                 merge_five_cell_four, merge_five_cell_five) is True:
                                merge_one_row(merge_five_sheet, merge_five_cell_one)
                                return [merge_five_cell_two, merge_five_cell_one]
                        if len(value[l].get("Cell")) < 2:
                            first_cell = split_cell_value(value[l], 0)

                            merge_one_row(self.sheet, first_cell)
                            inset_cell_value(self.sheet, value, l, first_cell, 0, 1)
                            self.color_cell(value[l].get("Department"), first_cell[0] + first_cell[1])
                        # If course happens twice a week
                        elif len(value[l].get("Cell")) == 2:
                            first_cell = split_cell_value(value[l], 0)
                            second_cell = split_cell_value(value[l], 1)
                            # Merging courses if they have same time and same room
                            cell = merge_two_rows(self.sheet, first_cell, second_cell)
                            if cell is True:
                                inset_cell_value(self.sheet, value, l, first_cell, 0, 1)
                                # Colors the cell by specific color
                                self.color_cell(value[l].get("Department"), first_cell[0] + first_cell[1])
                            if cell is not True:
                                for c in cell:
                                    inset_cell_value(self.sheet, value, l, c, 0, 1)
                                    self.color_cell(value[l].get("Department"), c[0] + c[1])

                                if value[l].get("Time_Comment") is None:
                                    self.sheet[first_cell[0] + first_cell[1]] = value[l].get("Course")
                                    self.sheet[second_cell[0] + second_cell[1]] = value[l].get("Course")
                                    self.color_cell(value[l].get("Department"), first_cell[0] + first_cell[1])
                                    self.color_cell(value[l].get("Department"), second_cell[0] + second_cell[1])
                                else:
                                    self.sheet[first_cell[0] + first_cell[1]] = value[l].get("Course") + "\n" + \
                                                                                value[l].get("Time_Comment")
                                    self.sheet[second_cell[0] + second_cell[1]] = \
                                        value[l].get("Course") + "\n" + value[l].get("Time_Comment")
                                    self.color_cell(value[l].get("Department"), first_cell[0] + first_cell[1])
                                    self.color_cell(value[l].get("Department"), second_cell[0] + second_cell[1])

                        elif len(value[l].get("Cell")) == 3:
                            first_cell = split_cell_value(value[l], 0)
                            second_cell = split_cell_value(value[l], 1)
                            third_cell = split_cell_value(value[l], 2)

                            # Merging courses if they have same time and same room
                            cell = merge_three_rows(self.sheet, first_cell, second_cell, third_cell)
                            if cell is True:
                                inset_cell_value(self.sheet, value, l, first_cell, 0, 1)
                                # Colors the cell by specific color
                                self.color_cell(value[l].get("Department"), first_cell[0] + first_cell[1])
                            if cell is not True:
                                for c in cell:
                                    inset_cell_value(self.sheet, value, l, c, 0, 1)
                                    self.color_cell(value[l].get("Department"), c[0] + c[1])

                        elif len(value[l].get("Cell")) == 4:
                            first_cell = split_cell_value(value[l], 0)
                            second_cell = split_cell_value(value[l], 1)
                            third_cell = split_cell_value(value[l], 2)
                            fourth_cell = split_cell_value(value[l], 3)

                            cell = merge_four_rows(self.sheet, first_cell, second_cell, third_cell, fourth_cell)
                            if cell is True:
                                inset_cell_value(self.sheet, value, l, first_cell, 0, 1)
                                # Colors the cell by specific color
                                self.color_cell(value[l].get("Department"), first_cell[0] + first_cell[1])
                            if cell is not True:
                                for c in cell:
                                    inset_cell_value(self.sheet, value, l, c, 0, 1)
                                    self.color_cell(value[l].get("Department"), c[0] + c[1])

                        elif len(value[l].get("Cell")) == 5:
                            first_cell = split_cell_value(value[l], 0)
                            second_cell = split_cell_value(value[l], 1)
                            third_cell = split_cell_value(value[l], 2)
                            fourth_cell = split_cell_value(value[l], 3)
                            fifth_cell = split_cell_value(value[l], 4)
                            cell = merge_five_rows(self.sheet, first_cell, second_cell, third_cell, fourth_cell,
                                                   fifth_cell)
                            if cell is True:
                                inset_cell_value(self.sheet, value, l, first_cell, 0, 1)
                                # Colors the cell by specific color
                                self.color_cell(value[l].get("Department"), first_cell[0] + first_cell[1])
                            if cell is not True:
                                for c in cell:
                                    inset_cell_value(self.sheet, value, l, c, 0, 1)
                                    self.color_cell(value[l].get("Department"), c[0] + c[1])

        self.border_all_cells("A1")

    def color_cell(self, course_department, coordinate, course_type_list=True):
        """Colors a course based on a department color"""
        # Different colors for each department
        # Accounting
        acct = PatternFill(start_color='FF958C', end_color='FF958C',
                           fill_type='solid')
        # Business Law
        blaw = PatternFill(start_color='FFCC00', end_color='FFCC00',
                           fill_type='solid')
        # Business
        bus = PatternFill(start_color='FFFF00', end_color='FFFF00',
                          fill_type='solid')
        # Finance
        fin = PatternFill(start_color='99CC00', end_color='99CC00',
                          fill_type='solid')
        # International Business
        ibus = PatternFill(start_color='8CF6FF', end_color='8CF6FF',
                           fill_type='solid')
        # Master of Business Administration
        mba = PatternFill(start_color='33CCCC', end_color='33CCCC',
                          fill_type='solid')
        # Master of Accounting
        macc = PatternFill(start_color='FF00FF', end_color='FF00FF',
                           fill_type='solid')
        # Management
        mgmt = PatternFill(start_color='CC99FF', end_color='CC99FF',
                           fill_type='solid')
        # Marketing
        mrkt = PatternFill(start_color='A28CFF', end_color='A28CFF',
                           fill_type='solid')

        # Checks for the first two letters to identify the color
        if course_department == "Accounting":
            color = acct
            if course_type_list is True:
                self.course_types_list.append("Accounting")
        # Business Law
        elif course_department == "Business Law":
            color = blaw
            if course_type_list is True:
                self.course_types_list.append("Business Law")
        # Business
        elif course_department == "Business":
            color = bus
            if course_type_list is True:
                self.course_types_list.append("Business")
        # Finance
        elif course_department == "Finance":
            color = fin
            if course_type_list is True:
                self.course_types_list.append("Finance")
        # International Business
        elif course_department == "International Business":
            color = ibus
            if course_type_list is True:
                self.course_types_list.append("International Business")
        # Master of Business Administration
        elif course_department == "MBA":
            color = mba
            if course_type_list is True:
                self.course_types_list.append("MBA")
        # Master of Accounting
        elif course_department == "MACC":
            color = macc
            if course_type_list is True:
                self.course_types_list.append("MACC")
        # Management
        elif course_department == "Management":
            color = mgmt
            if course_type_list is True:
                self.course_types_list.append("Management")
        # Marketing
        elif course_department == "Marketing":
            color = mrkt
            if course_type_list is True:
                self.course_types_list.append("Marketing")
        else:
            color = PatternFill(start_color='EEEFEF', end_color='EEEFEF',
                                fill_type='solid')
        # Fills the color
        self.sheet[coordinate].fill = color
        # Makes the text be in the center of a cell
        self.sheet[coordinate].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    def color_cell_meaning(self, row_num):
        """Will appear on a right side which color assign to which department"""
        get_max_column = self.sheet.max_column
        row = row_num

        def remove_duplicates(course_list):
            return list(set(course_list))

        def sort_list(course_list):
            course_list.sort()
            return course_list

        unique_types = sort_list(remove_duplicates(self.course_types_list))

        for i in range(len(unique_types)):
            alphabet = ''.join(string.ascii_uppercase[get_max_column+1])
            self.color_cell(unique_types[i], alphabet+str(row), False)
            self.sheet[''.join(string.ascii_uppercase[get_max_column+2])+str(row)] = "-" + unique_types[i]
            row += 1

    def set_page_break(self):
        # 40 rows per page
        get_max_row = self.sheet.max_row
        get_max_column = self.sheet.max_column

        if len(self.days) == 4:
            if get_max_row >= 40:
                while get_max_row >= 40:
                    self.sheet.sheet_properties.pageSetUpPr.fitToPage = True
                    openpyxl.worksheet.pagebreak.PageBreak.tagname = 'rowBreaks'
                    page_break_row = Break((get_max_row + 1)-37)
                    self.sheet.page_breaks.append(page_break_row)

                    openpyxl.worksheet.pagebreak.PageBreak.tagname = 'colBreaks'
                    page_break_column = Break(get_max_column + 1)
                    self.sheet.page_breaks.append(page_break_column)
                    get_max_row -= 37

            elif get_max_row == 40:
                pass
            else:
                self.sheet.sheet_properties.pageSetUpPr.fitToPage = True
                openpyxl.worksheet.pagebreak.PageBreak.tagname = 'rowBreaks'
                page_break_row = Break(get_max_row + 1)
                self.sheet.page_breaks.append(page_break_row)

                openpyxl.worksheet.pagebreak.PageBreak.tagname = 'colBreaks'
                page_break_column = Break(get_max_column + 1)
                self.sheet.page_breaks.append(page_break_column)
        # Landscape orientation
        self.sheet.page_setup.orientation = self.sheet.ORIENTATION_LANDSCAPE
