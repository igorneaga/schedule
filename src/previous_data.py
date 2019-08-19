import datetime
import os
import string
import urllib.parse

import openpyxl
import requests
from bs4 import BeautifulSoup
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.styles.borders import Border, Side
from openpyxl.worksheet.pagebreak import Break

from src.previous_semesters import ReceiveSemesters


class PreviousCourses:
    def __init__(self, department, semester, year, semester_parameters=None, department_parameters=None):
        self.user_selected_department = department
        self.user_selected_year = year
        self.user_selected_semester = semester
        self.department_parameters = department_parameters
        self.semester_parameters = semester_parameters

        self.request_headers = {
                                'Content-Type': "application/x-www-form-urlencoded",
                                'Origin': "https://secure2.mnsu.edu",
                                'Referer': "https://secure2.mnsu.edu/courses/Default.asp",
                                'cache-control': "no-cache",
                                'Postman-Token': "6f1fa71c-c6fa-4fc4-b7df-e3cefe723179"
                                }
        self.user_request_encode = None

        self.course_list = []
        self.main_class_controller()

    def main_class_controller(self):
        def transfer_department_name(department_abbreviation):
            """Transfers to the full department name"""
            return {
                    'ACCT': 'Accounting',
                    'BLAW': 'Business Law',
                    'FINA': 'Finance',
                    'IBUS': 'International Business',
                    'MGMT': 'Management',
                    'MRKT': 'Marketing',
                    'MACC': 'Master in Accounting'
                    }.get(department_abbreviation, 'Accounting')

        def get_payload_encode(encode_params, url, year, semester, department):
            """Gets parameters for semester and subject"""
            university_courses_url = url
            page_response = requests.get(university_courses_url)
            url_html_parser = BeautifulSoup(page_response.content, "html.parser")

            semester_option = semester + str(year)
            semester_option = semester_option.upper()

            department_option = (department.replace(" ", "")).upper()

            for option in url_html_parser.find_all('option'):
                search_option = (option.text.replace(" ", "")).upper()
                if semester_option == search_option:
                    encode_params['semester'] = option['value']
                if search_option[0:len(department_option)] == department_option:
                    encode_params['subject'] = option['value']

            return encode_params, semester_option

        def transfer_params(parse_params):
            """Urlparse"""
            parse_params = urllib.parse.urlencode(parse_params)

            params_list = [parse_params]
            return params_list

        full_department_name = transfer_department_name(department_abbreviation=self.user_selected_department)
        params = {
            'semester': None,
            'campus': '1,2,3,4,5,6,7,9,A,B,C,I,L,M,N,P,Q,R,S,T,W,U,V,X,Y,Z',
            'startTime': '0600',
            'endTime': '2359',
            'days': 'ALL',
            'All': 'All Sections',
            'subject': None,
            'undefined': ''
        }
        # if not None, skips one function to increase performance
        if self.department_parameters is not None:
            params['subject'] = self.department_parameters
            params['semester'] = self.semester_parameters
            self.user_request_encode = transfer_params(params)
        else:
            web_params, semester_option = get_payload_encode(params, ReceiveSemesters.COURSES_URL, self.user_selected_year, self.user_selected_semester, full_department_name)
            self.user_request_encode = transfer_params(web_params)

        response = requests.request("POST", ReceiveSemesters.COURSES_URL, data=self.user_request_encode[0], headers=self.request_headers)
        self.get_data(response)
        CreateStandardTable(self.course_list, full_department_name, self.user_selected_semester, str(self.user_selected_year), self.user_selected_department)

    def get_data(self, web_response):
        data_html_parser = BeautifulSoup(web_response.text, 'html.parser')

        for table_data in data_html_parser.find_all("tr"):
            course_titles_list = []
            course_data_list = []

            course_title_raw = table_data.find(color="#ffffff")
            if course_title_raw is not None:
                for course_title in course_title_raw("b"):
                    title_text = course_title.get_text()
                    course_titles_list.append(title_text)
            if (table_data["bgcolor"] == "#E1E1CC") or (table_data["bgcolor"] == "#FFFFFF"):
                for course_data in table_data("td"):
                    course_data_text = course_data.get_text()
                    course_data_list.append(course_data_text)

            if course_titles_list:
                self.course_list.append(course_titles_list)
            if course_data_list:
                if len(course_data_list[0]) == 6:
                    self.course_list.append(course_data_list)


class CreateStandardTable:
    def __init__(self, web_course_data, departament_full, semester, year, department_abbreviation):
        self.full_departament_name = departament_full
        self.department_abbreviation = department_abbreviation
        self.semester = semester
        self.year = year

        self.web_course_data = web_course_data

        self.excel_workbook = None
        self.excel_sheet = None

        self.main_program_controller()

    def main_program_controller(self):
        self.create_excel_file()

        self.create_excel_sheet()

        self.department_heading()
        self.excel_column_headings()
        self.set_data()
        self.adjust_cells_width()
        self.border_all_cells("A1")
        self.set_page_break()
        self.excel_workbook.save('web_files\\' + self.department_abbreviation.replace(" ", "_")[0:27] + "_" +
                                 self.year + ".xlsx")

    def create_excel_file(self):
        def create_directory():
            """Creates directory for created excel file"""
            if not os.path.exists('web_files'):
                os.makedirs('web_files')
        create_directory()

        self.excel_workbook = openpyxl.Workbook()

    def create_excel_sheet(self):
        today_year = datetime.datetime.today().strftime('%Y')

        self.excel_sheet = self.excel_workbook.get_sheet_by_name("Sheet")
        self.excel_sheet.title = self.full_departament_name + "_" + today_year

    def department_heading(self):
        self.excel_sheet.merge_cells("A1:M2")

        self.excel_sheet["A1"] = self.full_departament_name.upper() + " " + self.semester.upper() + " " + self.year
        self.excel_sheet["A1"].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        self.excel_sheet["A1"].font = Font(sz=12, bold=True, italic=False)

        cell_color = PatternFill(start_color='EEEEEE', end_color='EEEEEE', fill_type='solid')
        self.excel_sheet["A1"].fill = cell_color

    def excel_column_headings(self):
        excel_headings_list = ["Course", "Number", "Section", "Credits", "Title of Course", "Room", "Days", "Time",
                               "Enrollment", "Faculty", "Start Date", "End Date", "Notes"]
        for heading_column in range(len(excel_headings_list)):
            heading_coordinate = ''.join(string.ascii_uppercase[heading_column]) + "3"  # i.e "A3"
            self.excel_sheet[heading_coordinate] = excel_headings_list[heading_column]
            self.excel_sheet[heading_coordinate].alignment = Alignment(horizontal='center', vertical='center',
                                                                       wrap_text=True)
            self.excel_sheet[heading_coordinate].font = Font(sz=12, bold=False, italic=False)

    def set_data(self):
        starting_excel_row = 4
        excel_cell_index = 0

        def insert_cells(web_courses_data, excel_sheet, cell_index, starting_row):
            web_courses_data = web_courses_data
            excel_sheet = excel_sheet
            cell_index = cell_index
            starting_row = starting_row
            if web_courses_data[cell_index][0]:
                if web_courses_data[cell_index][0].isdigit() is not True:
                    course = web_courses_data[cell_index][0]
                    course_number = web_courses_data[cell_index][1]
                    symbol_index = web_courses_data[cell_index][3].find("(")
                    course_credits = str(web_courses_data[cell_index][3][symbol_index + 1:symbol_index + 2])
                    course_title = str(web_courses_data[cell_index][2]).strip()

                    while len(web_courses_data[cell_index + 1][0]) == 6 and (
                            web_courses_data[cell_index + 1][0].isdigit()):
                        excel_sheet["A" + str(starting_row)] = course
                        excel_sheet["B" + str(starting_row)] = int(course_number)
                        excel_sheet["C" + str(starting_row)] = int(web_courses_data[cell_index + 1][1][0:2])
                        excel_sheet["D" + str(starting_row)] = int(course_credits)
                        excel_sheet["E" + str(starting_row)] = course_title

                        if web_courses_data[cell_index + 1][4][0:3] == "ARR":
                            excel_sheet["F" + str(starting_row)] = "ARR"
                        else:
                            excel_sheet["F" + str(starting_row)] = web_courses_data[cell_index + 1][6]
                        excel_sheet["G" + str(starting_row)] = str(web_courses_data[cell_index + 1][3]).strip()

                        if web_courses_data[cell_index + 1][6].replace(" ", "") == "ONLINE":
                            excel_sheet["F" + str(starting_row)] = "ONLINE"
                            excel_sheet["H" + str(starting_row)] = "ONLINE"
                        else:
                            excel_sheet["H" + str(starting_row)] = web_courses_data[cell_index + 1][4]

                        excel_sheet["I" + str(starting_row)] = int(web_courses_data[cell_index + 1][9])
                        excel_sheet["J" + str(starting_row)] = web_courses_data[cell_index + 1][7]

                        start_end_date = web_courses_data[cell_index + 1][5].split("-")
                        start_end_date[0] = start_end_date[0].replace(" ", "")
                        start_end_date[0] = start_end_date[0][:-2] + '20' + start_end_date[0][-2:]
                        start_end_date[1] = start_end_date[1].replace(" ", "")
                        start_end_date[1] = start_end_date[1][:-2] + '20' + start_end_date[1][-2:]

                        excel_sheet["K" + str(starting_row)] = start_end_date[0]
                        excel_sheet["L" + str(starting_row)] = start_end_date[1]

                        for all_columns in range(12):
                            cell_coordinate = ''.join(string.ascii_uppercase[all_columns]) + str(starting_row)
                            self.excel_sheet[cell_coordinate].alignment = Alignment(horizontal='center',
                                                                                    vertical='center',
                                                                                    wrap_text=True)

                        cell_index += 1
                        starting_row += 1
                    else:
                        insert_cells(web_courses_data, excel_sheet, cell_index + 1, starting_row)

        try:
            insert_cells(self.web_course_data, self.excel_sheet, excel_cell_index, starting_excel_row)
        except IndexError:
            pass

    def adjust_cells_width(self):
        """Adjusts all the cell width. It is really cool"""

        for column in self.excel_sheet.columns:
            max_cell_length = 0
            get_column = column[0].column

            if get_column is "A":
                self.excel_sheet.column_dimensions["A"].width = 9
            elif get_column is "B":
                self.excel_sheet.column_dimensions["A"].width = 7.9
            else:
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_cell_length:
                            max_cell_length = len(cell.value)
                    except TypeError:
                        pass

                # A formula for auto adjusted width
                adjusted_width = (max_cell_length + 2) * 1.02

                # Limit width
                if adjusted_width > 25:
                    adjusted_width = 25

                self.excel_sheet.column_dimensions[get_column].width = adjusted_width

    def border_all_cells(self, start_cell):
        """Borders all table"""
        # Gets table size
        excel_max_row = self.excel_sheet.max_row
        excel_max_column = self.excel_sheet.max_column

        # Style of a border
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))

        # Goes over each cell and applies border
        for column in range(excel_max_column):
            # Transfers column to an alphabetical format
            col_letter = ''.join(string.ascii_uppercase[column])
            for row in range(excel_max_row):
                row += 1
                self.excel_sheet[col_letter + str(row)].border = thin_border

    def set_page_break(self):
        # 40 rows per page
        get_max_row = self.excel_sheet.max_row
        get_max_column = self.excel_sheet.max_column

        if get_max_row >= 32:
            while get_max_row >= 32:
                self.excel_sheet.sheet_properties.pageSetUpPr.fitToPage = True
                openpyxl.worksheet.pagebreak.PageBreak.tagname = 'rowBreaks'
                page_break_row = Break((get_max_row + 1) - 31)
                self.excel_sheet.page_breaks.append(page_break_row)

                openpyxl.worksheet.pagebreak.PageBreak.tagname = 'colBreaks'
                page_break_column = Break(get_max_column + 1)
                self.excel_sheet.page_breaks.append(page_break_column)
                get_max_row -= 32
        else:
            self.excel_sheet.sheet_properties.pageSetUpPr.fitToPage = True
            openpyxl.worksheet.pagebreak.PageBreak.tagname = 'rowBreaks'
            page_break_row = Break(get_max_row + 1)
            self.excel_sheet.page_breaks.append(page_break_row)

            openpyxl.worksheet.pagebreak.PageBreak.tagname = 'colBreaks'
            page_break_column = Break(get_max_column + 1)
            self.excel_sheet.page_breaks.append(page_break_column)
        # Landscape orientation
        self.excel_sheet.page_setup.orientation = self.excel_sheet.ORIENTATION_LANDSCAPE
