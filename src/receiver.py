import datetime
import math
import os
import re
from itertools import chain  # For more efficient looping
from tkinter import messagebox

import openpyxl
from openpyxl.comments import Comment
from openpyxl.styles import PatternFill

from src import room_schedule_table, payroll_table_process


class DataProcessor:

    def __init__(self, folder, file_directory, table_name, table_semester, table_year, table_type, friday,
                 classroom_capacity, payroll):
        self.save_path = folder
        self.file_directory = file_directory
        self.table_name = table_name
        self.table_semester = table_semester
        self.table_year = table_year
        self.days_order = table_type
        self.friday_choice = friday
        self.classroom_capacity = classroom_capacity

        self.payroll = payroll

        self.course_days = []
        self.excel_data_list = None
        self.dict_courses_list = []
        self.different_date_list = []

        self.file_paths_list = []
        self.user_excel_errors = []

        # User-Excel-Columns identification
        self.excel_course_name = 1
        self.excel_course_number = 2
        self.excel_course_section = 3
        self.excel_course_credits = 4
        self.excel_course_title = 5
        self.excel_course_room = 6
        self.excel_course_days = 7
        self.excel_course_time = 8

        # Start-user-data
        self.starting_excel_row = 4

        # In case of any errors, it will give three chances before the program stops
        self.number_close_trials = 0

        self.main_class_order()

    def main_class_order(self):
        """Main program logic"""

        def assign_days_order(user_choice):
            """Returns a day's order selected by the user"""
            # 1 - out order days
            # 2 - in order days
            if user_choice is 2:
                return ["Monday", "Tuesday", "Wednesday", "Thursday"]
            return ["Monday", "Wednesday", "Tuesday", "Thursday"]

        def friday_option(user_friday_option, current_days):
            """Checks if the user selected the Friday option. Returns it in the day's list"""
            if user_friday_option == 1:
                current_days.append("Friday")
                return current_days
            return current_days

        def create_excel_copies():
            """Creates a folder to store all the copy files"""
            if not os.path.exists('copy_folder'):
                # Creates a folder for files
                os.makedirs('copy_folder')

        def close_file_error():
            user_response = messagebox.showerror("Close File", "Save and close excel files to continue... \n\n"
                                                 + str(permission_error_message))
            return user_response

        try:
            table_order = assign_days_order(self.days_order)
            self.course_days = friday_option(self.friday_choice, table_order)
            create_excel_copies()
            for i in range(len(self.file_directory)):
                get_file_name = os.path.basename(self.file_directory[i])
                workbook_copy = openpyxl.load_workbook(self.file_directory[i])
                set_file_path = 'copy_folder\\' + 'copy_' + get_file_name
                self.get_excel_data(workbook_copy, self.file_directory[i])
                self.set_dict_courses(self.excel_data_list)
                if self.payroll is False:
                    self.time_conflict()
                    self.file_paths_list.append(set_file_path)
                    self.color_comment_copy_excel(workbook_copy, self.file_directory[i])
                    workbook_copy.save(set_file_path)
            self.get_excel_errors()
            if self.payroll is False:
                self.create_excel_table()
            else:
                payroll_table_process.PayrollTable(self.dict_courses_list, self.save_path)

        except PermissionError as permission_error_message:
            # Gives a user three chances to close excel files
            if self.number_close_trials > 2:
                self.user_excel_errors = 'User_Doesnt_Listen'
            else:
                msg_response = close_file_error()
                if msg_response == "ok":
                    self.number_close_trials += 1
                    # Resets all the variables
                    self.excel_data_list = None
                    self.course_days = []
                    self.user_excel_errors = []
                    self.file_paths_list = []
                    self.dict_courses_list = []
                    self.different_date_list = []
                    self.main_class_order()

    def get_excel_data(self, wb_copy, file_path):
        """Gets all information from excel file"""
        worksheet_excel_data = []
        for number_of_files in range(len(wb_copy.worksheets)):
            get_excel_workbook = openpyxl.load_workbook(file_path, read_only=True)

            get_sheet_name = get_excel_workbook.get_sheet_names()[number_of_files]
            worksheet_read_mode = get_excel_workbook[get_sheet_name]
            worksheet_excel_data.append(list(self.iter_rows(worksheet_read_mode, file_path, get_sheet_name)))
        self.excel_data_list = (list(chain.from_iterable(worksheet_excel_data)))

    def create_report_dictionary(self, excel_row, excel_column, excel_file_path, excel_sheet_name, error_color,
                                 error_comment=None):
        """If there is any conflict, error, missing data it will store that information in user_excel_errors
        dictionary"""

        report_dict = {
            'Row': excel_row,
            'Column': excel_column,
            'File_Path': excel_file_path,
            'Sheet_Name': excel_sheet_name,
            'Color': error_color
        }

        if error_comment is None:
            pass
        else:
            report_dict["Comment"] = error_comment

        self.user_excel_errors.append(report_dict.copy())

    def iter_rows(self, read_worksheet, file_path, sheet_name):
        """Loops through rows. Checks if the value is none. And transforms time to the same format"""

        def convert_time(user_time):
            """Converts time into the correct format"""
            try:
                user_time = datetime.datetime.strptime(user_time, "%H:%M")
                converted_time = datetime.datetime.strftime(user_time, '%I:%M')
                return converted_time
            except ValueError:
                try:
                    user_time = datetime.datetime.strptime(user_time, "%H")
                    converted_time = datetime.datetime.strftime(user_time, '%I:%M')
                    return converted_time
                except ValueError:
                    try:
                        user_time = datetime.datetime.strptime(user_time, "%H:%M %p")
                        converted_time = datetime.datetime.strftime(user_time, '%I:%M')
                        return converted_time
                    except ValueError:
                        try:
                            user_time = datetime.datetime.strptime(user_time, "%H:%M%p")
                            converted_time = datetime.datetime.strftime(user_time, '%I:%M')
                            return converted_time
                        except ValueError:
                            pass

        data_start_row = self.starting_excel_row
        iterable_row = read_worksheet.iter_rows()

        # Skips unnecessary rows
        for skip_row in range(data_start_row - 1):
            next(iterable_row)

        for row in iterable_row:
            # Creating temporary list
            excel_cell_value = []
            for cell in row:
                try:
                    if cell.value is None:
                        excel_cell_value.append(cell.value)
                    # Checks if course is Online
                    elif (cell.column == self.excel_course_time) & (cell.value in {"Online", "ONLINE", "online"}):
                        excel_cell_value.append("Online-Online")
                    elif cell.column == self.excel_course_time:
                        try:

                            split_time = cell.value.split('-')

                            split_time[0] = "".join(split_time[0].split())
                            split_time[1] = "".join(split_time[1].split())

                            # Converts our time to the same format
                            converted_start_time = convert_time(split_time[0])
                            converted_end_time = convert_time(split_time[1])

                            excel_cell_value.append(converted_start_time + "-" + converted_end_time)
                        except AttributeError:
                            # Will mark if program can't read the datetime
                            if type(cell.value) is datetime.time:
                                excel_cell_value.append('%s:%s-None' % (cell.value.hour, cell.value.minute))
                                # Will notify user about possible issue
                                comment = "Does it contain the start time or end time of the course?"
                                self.create_report_dictionary(cell.row, cell.column, file_path, sheet_name,
                                                              error_color="FF687B", error_comment=comment)
                            elif cell.value is not None:

                                comment = "Does it contain the start time or end time of the course?"
                                self.create_report_dictionary(cell.row, cell.column, file_path, sheet_name,
                                                              error_color="FF687B", error_comment=comment)

                            else:
                                excel_cell_value.append("None-None")
                        except IndexError:
                            excel_cell_value.append(cell.value)
                    else:
                        excel_cell_value.append(cell.value)
                except AttributeError:
                    pass

            if all(none_values is None for none_values in excel_cell_value):
                data_start_row += 1
            else:
                excel_cell_value.insert(0, file_path)
                excel_cell_value.insert(0, sheet_name)
                excel_cell_value.insert(0, data_start_row)

                data_start_row += 1
            yield excel_cell_value

    def color_comment_copy_excel(self, workbook, file_path):
        """Will comment and color excel copy sheets where the program finds any mistakes or conflicts.
                It will help the user to find them faster than ever."""

        get_excel_workbook = workbook
        # Goes through each error

        for len_errors in range(len(self.user_excel_errors)):
            # Checks if this error is assign to this file
            if self.user_excel_errors[len_errors].get("File_Path") == file_path:
                self.user_excel_errors[len_errors].get("File_Path")
                edit_sheet = get_excel_workbook[self.user_excel_errors[len_errors].get("Sheet_Name")]
                # Takes error information
                row_num = self.user_excel_errors[len_errors].get("Row")
                column_num = self.user_excel_errors[len_errors].get("Column")
                error_color = self.user_excel_errors[len_errors].get("Color")
                # Fills cell with error color
                edit_sheet.cell(row=row_num, column=column_num).fill = PatternFill(start_color=error_color,
                                                                                   end_color=error_color,
                                                                                   fill_type='solid')
                # Comments cell
                if self.user_excel_errors[len_errors].get("Comment") is not None:
                    comment = self.user_excel_errors[len_errors].get("Comment")
                    edit_sheet.cell(row=row_num, column=column_num).comment = Comment(comment, author="TableMaker")

    def mark_none_values(self, data_list):
        """Marks none values with specific color"""
        for len_data in range(len(data_list)):
            # Goes every None value
            none_index = [i for i, e in enumerate(data_list[len_data]) if e == "None"]
            for val_length in range(len(none_index)):
                column = none_index[val_length] - 2
                # if column is above 13 we don't need to mark
                if column > 13:
                    pass
                else:
                    try:
                        self.create_report_dictionary(data_list[len_data][0], column, data_list[len_data][2],
                                                      data_list[len_data][1], error_color="FADDA7")
                    except IndexError:
                        pass

    def set_dict_courses(self, excel_data):
        """Inserting all information into a dictionary"""

        def clear_unnecessary_list(data_list):
            data_list = list(filter(None, data_list))  # Deletes empty lists

            # Removes a list which contains only None values
            clearing_none_list = []
            for list_number in data_list:
                if all(per_list is None
                       for per_list in list_number):
                    pass
                else:
                    clearing_none_list.append(list_number)
            # Erase previous list
            data_list = []
            # Makes None vales into String ('None')
            for none_list in clearing_none_list:
                temp_list = []
                for element in none_list:
                    if element is None:
                        temp_list.append('None')
                    else:
                        temp_list.append(element)
                data_list.append(temp_list)

            return data_list

        def course_room_format(room_number):
            """Formats room to follow the same format"""
            room_number = room_number.replace(" ", "")
            str_list = re.split('(\\d+)', room_number)
            # Removes empty str
            filter_str_list = list(filter(None, str_list))
            for f in range(len(filter_str_list)):
                filter_str_list[f] = filter_str_list[f].lower()
                filter_str_list[f] = ''.join(filter_str_list[f].split())
                filter_str_list[f] = filter_str_list[f].upper()

                if f == 1:
                    while len(filter_str_list[f]) < 4:
                        filter_str_list[f] = '0' + filter_str_list[f]

            return ' '.join(filter_str_list)

        def course_title_format(course_title, course_number, course_section):
            """Setting course title to the same format"""
            c_title = course_title.replace(' ', '')
            c_number = course_number.replace(' ', '')
            c_section = course_section.replace(' ', '')
            return f'{c_title.upper()} {c_number}-{c_section}'

        def convert_user_days(day, k):
            """Transfers day into proper format"""
            # Transforms to Monday
            if any(c in day[k].upper() for c in "M") \
                    or any(c in day[k:3].upper() for c in "MON") \
                    or any(c in day[k:6].upper() for c in "MONDAY"):
                return "Monday"

            # Transforms to Thursday
            elif any(c in day[k].upper() for c in ("R", "H")) \
                    or day[k:2].upper() == "TH" \
                    or day[k:3].upper() == "THU" \
                    or day[k:7].upper() == "THURSDAY":
                return "Thursday"

            # Transforms to Tuesday
            elif any(c in day[k].upper() for c in "T") \
                    or any(c in day[k:3].upper() for c in "TUE") \
                    or any(c in day[k:7].upper() for c in "TUESDAY"):
                return "Tuesday"

            # Transforms to Wednesday
            elif any(c in day[k].upper() for c in "W") \
                    or any(c in day[k:3].upper() for c in "WED") \
                    or any(c in day[k:7].upper() for c in "WEDNESDAY"):
                return "Wednesday"

            # Transforms to Friday
            elif any(c in day[k].upper() for c in "F") \
                    or any(c in day[k:3].upper() for c in "FRI") \
                    or any(c in day[k:7].upper() for c in "FRIDAY"):
                return "Friday"

            # By default
            return 'None'

        excel_data = clear_unnecessary_list(excel_data)
        self.mark_none_values(excel_data)

        def set_online_course(d_courses, l_data):
            """Sets online courses to a specific type."""
            d_courses["Course_Days"] = []
            d_courses["Start_Time"] = "Online"
            d_courses["End_Time"] = "Online"
            d_courses["Type"] = ["Online"]
            d_courses["Row"] = l_data[0]
            d_courses["File"] = l_data[2]
            d_courses["Sheet_Name"] = l_data[1]
            return d_courses

        def set_course_type(d_courses):
            """Sets a course type if follows specific conditions"""
            d_courses["Type"] = ["Classroom"]
            if d_courses.get("Course")[0:3] == "MBA":
                d_courses.get("Type").append("MBA")
            elif d_courses.get("Course")[0:4] == "MACC":
                d_courses.get("Type").append("MACC")
            if d_courses.get("Room")[0:2] == "HU":
                d_courses.get("Type").append("Hubbard")
            if d_courses.get("Room")[0:7] == "MH 0211":
                d_courses.get("Type").append("Telepresence")
            return d_courses.get("Type")

        def set_course_department(course_title):
            return {
                'AC': 'Accounting',
                'BL': 'Business Law',
                'BU': 'Business',
                'FI': 'Finance',
                'IB': 'International Business',
                'MB': 'MBA',  # Master of Business Administration
                'MA': 'MACC',  # Master of Accounting
                'MG': 'Management',
                'MR': 'Marketing'
            }.get(course_title[:2], None)

        for each_excel_data in excel_data:
            try:
                dict_courses = dict()
                # Checks the room column contains an online word
                if course_room_format(str(each_excel_data[self.excel_course_room + 2])).upper() == "ONLINE":
                    dict_courses["Course"] = course_title_format(str(each_excel_data[self.excel_course_name + 2]),
                                                                 str(each_excel_data[self.excel_course_number + 2]),
                                                                 str(each_excel_data[self.excel_course_section + 2]))

                    dict_courses["Credits"] = each_excel_data[6]
                    dict_courses["Course_Title"] = each_excel_data[7]
                    dict_courses["Faculty"] = each_excel_data[12]
                    dict_courses["Enrollment"] = each_excel_data[11]
                    dict_courses = set_online_course(dict_courses, each_excel_data)

                else:
                    time_split = each_excel_data[self.excel_course_time + 2].split('-')
                    # Sets course to specific format
                    dict_courses["Course"] = course_title_format(str(each_excel_data[self.excel_course_name + 2]),
                                                                 str(each_excel_data[self.excel_course_number + 2]),
                                                                 str(each_excel_data[self.excel_course_section + 2]))
                    # Sets room to specific format.
                    dict_courses["Room"] = course_room_format(str(each_excel_data[self.excel_course_room + 2]))

                    # Checks if the time contains Online word
                    if time_split[0].upper() == "ONLINE":
                        dict_courses["Room"] = course_room_format(str(each_excel_data[self.excel_course_room + 2]))
                        dict_courses = set_online_course(dict_courses, each_excel_data)
                    else:
                        dict_courses["Room"] = course_room_format(str(each_excel_data[self.excel_course_room + 2]))
                        dict_courses["Course_Days"] = []

                        dict_courses["Row"] = each_excel_data[0]
                        dict_courses["File"] = each_excel_data[2]
                        dict_courses["Sheet_Name"] = each_excel_data[1]
                        dict_courses["Credits"] = each_excel_data[6]
                        dict_courses["Course_Title"] = each_excel_data[7]
                        dict_courses["Type"] = set_course_type(dict_courses)
                        dict_courses["Course_Title"] = each_excel_data[7]
                        dict_courses["Enrollment"] = each_excel_data[11]
                        dict_courses["Faculty"] = each_excel_data[12]
                        try:
                            dict_courses["Start_Time"] = time_split[0]
                            dict_courses["End_Time"] = time_split[1]

                            # Due to space table concern, we need to limit evening classes cells
                            if time_split[0][1] == '6':
                                dict_courses["Time_Comment"] = " ends at " + time_split[1]
                                dict_courses["End_Time"] = "06:00"
                            if time_split[1][1] == '6' or time_split[1][1] == '7':
                                dict_courses["Time_Comment"] = " ends at " + time_split[1]
                                dict_courses["End_Time"] = "06:00"
                        except IndexError:
                            # Marks a course if program couldn't read it
                            dict_courses["Start_Time"] = dict_courses["End_Time"] = "None"
                            dict_courses["Type"] = "Error"

                        # Removes white spaces
                        if " " in each_excel_data[self.excel_course_days + 2]:
                            if "OR" in each_excel_data[self.excel_course_days + 2].upper():
                                modified_str = each_excel_data[self.excel_course_days + 2].upper()
                                each_excel_data[self.excel_course_days + 2] = modified_str.replace("OR", "")
                            if "AND" in each_excel_data[self.excel_course_days + 2].upper():
                                modified_str = each_excel_data[self.excel_course_days + 2].upper()
                                each_excel_data[self.excel_course_days + 2] = modified_str.replace("OR", "")

                            each_excel_data[self.excel_course_days + 2] = each_excel_data[
                                self.excel_course_days + 2].replace(" ", "")
                        # Splits days and converts to the proper format
                        if "," in each_excel_data[self.excel_course_days + 2]:
                            if "OR" in each_excel_data[self.excel_course_days + 2].upper():
                                modified_str = each_excel_data[self.excel_course_days + 2].upper()
                                each_excel_data[self.excel_course_days + 2] = modified_str.replace("OR", "")

                            if "AND" in each_excel_data[self.excel_course_days + 2].upper():
                                modified_str = each_excel_data[self.excel_course_days + 2].upper()
                                each_excel_data[self.excel_course_days + 2] = modified_str.replace("OR", "")

                            split_by_comma = [x.strip() for x in each_excel_data[self.excel_course_days + 2].split(',')]
                            for position in range(len(split_by_comma)):
                                # Checks if the function can convert user day format
                                if convert_user_days(split_by_comma[position], 0) == 'None':
                                    comment = dict_courses.get("Course") + " must follow the day format" + (' ' * 150)
                                    self.create_report_dictionary(dict_courses.get("Row"), 7, dict_courses.get("File"),
                                                                  dict_courses.get("Sheet_Name"), "FF687B", comment)
                                    dict_courses["Type"] = ["Error"]
                                dict_courses["Course_Days"].append(convert_user_days(split_by_comma[position], 0))

                        elif "None" in each_excel_data[self.excel_course_days + 2]:
                            pass

                        elif "ONLINE" in each_excel_data[self.excel_course_days + 2].upper():
                            dict_courses["Course"] = course_title_format(
                                str(each_excel_data[self.excel_course_name + 2]),
                                str(each_excel_data[self.excel_course_number + 2]),
                                str(each_excel_data[self.excel_course_section + 2]))
                            dict_courses = set_online_course(dict_courses, each_excel_data)

                        else:
                            for i in range(len(each_excel_data[self.excel_course_days + 2])):
                                # Checks if the function can convert user day format
                                if convert_user_days(each_excel_data[self.excel_course_days + 2], i) == 'None':
                                    comment = dict_courses.get("Course") + " must follow the day format" + (' ' * 150)
                                    self.create_report_dictionary(dict_courses.get("Row"), 7, dict_courses.get("File"),
                                                                  dict_courses.get("Sheet_Name"), "FF687B", comment)
                                    dict_courses["Type"] = ["Error"]

                                dict_courses["Course_Days"].append(
                                    convert_user_days(each_excel_data[self.excel_course_days + 2], i))

                if (each_excel_data[13] != "None") and (each_excel_data[13] is not None):
                    try:
                        if type(each_excel_data[13]) is str:
                            # Checking if the year is correct
                            if int(each_excel_data[13][-4:]) > (int(datetime.datetime.now().year) + 2) or int(
                                    each_excel_data[13][-4:]) < (int(datetime.datetime.now().year) - 2):
                                comment = dict_courses.get("Course") + " course might have a wrong year. " \
                                                                       "Please double check." + (' ' * 150)
                                self.create_report_dictionary(each_excel_data[0], 11, each_excel_data[2],
                                                              each_excel_data[1], 'FF687B', comment)

                            dict_courses["Start_Date"] = datetime.datetime.strptime(each_excel_data[13],
                                                                                    '%m/%d/%Y')
                        else:
                            dict_courses["Start_Date"] = each_excel_data[13]
                    except ValueError:
                        # If the date format is incorrect
                        comment = dict_courses.get("Course") + ' course does not match format "01/01/2020"' + \
                                  (' ' * 150)
                        self.create_report_dictionary(each_excel_data[0], 11, each_excel_data[2],
                                                      each_excel_data[1], 'FF687B', comment)

                if (each_excel_data[14] != "None") and (each_excel_data[13] is not None):
                    try:
                        if type(each_excel_data[14]) is str:
                            # Checking if the year is correct
                            if int(each_excel_data[14][-4:]) > (int(datetime.datetime.now().year) + 2) or \
                                    int(each_excel_data[14][-4:]) < (int(datetime.datetime.now().year) - 2):
                                comment = dict_courses.get("Course") + " course might have a wrong year. " \
                                                                       "Please double check." + (' ' * 150)
                                self.create_report_dictionary(each_excel_data[0], 12, each_excel_data[2],
                                                              each_excel_data[1], 'FF687B', comment)

                            dict_courses["End_Date"] = datetime.datetime.strptime(each_excel_data[14],
                                                                                  '%m/%d/%Y')
                        else:
                            dict_courses["End_Date"] = each_excel_data[14]
                    except ValueError:
                        # If the date format is incorrect
                        comment = dict_courses.get("Course") + ' course does not match format "01/01/2020"' + \
                                  (' ' * 150)
                        self.create_report_dictionary(each_excel_data[0], 12, each_excel_data[2],
                                                      each_excel_data[1], 'FF687B', comment)

                self.dict_courses_list.append(dict_courses.copy())
            except AttributeError:
                # If an error occurred, it will mark the whole row
                for i in range(12):
                    self.create_report_dictionary(each_excel_data[0], i + 1, each_excel_data[2], each_excel_data[1],
                                                  error_color="FF687B")
                # Marks last cell with comment
                comment = "A program couldn't read this row correctly. Report it if needed."
                self.create_report_dictionary(each_excel_data[0], 13, each_excel_data[2], each_excel_data[1],
                                              error_color="FF687B", error_comment=comment)

        def fixing_none_courses(all_courses):
            all_dict_courses = all_courses

            for len_dict in range(len(all_dict_courses)):
                for key in all_dict_courses[len_dict]:
                    if type(all_dict_courses[len_dict].get(key)) == str:
                        if all_dict_courses[len_dict].get(key)[0:4].upper() == 'NONE':
                            all_dict_courses[len_dict][key] = all_dict_courses[len_dict - 1].get(key)

                all_dict_courses[len_dict]["Department"] = set_course_department(all_dict_courses[len_dict].get("Course"))

        fixing_none_courses(self.dict_courses_list)

    def time_conflict(self):
        """Loops through each dictionary in the list. Looks for similar rooms and days.
        Finds a time conflict between courses. Deletes the conflict dict."""

        def check_course_dates(first_course, second_course):
            """Checks if courses has dates and dates differences"""

            # Dates from Minnesota State University, Mankato academic calendar
            now = datetime.datetime.now()
            first_year = now.year
            second_year = now.year + 1
            try:
                if first_course.year < second_course.year:
                    first_year = first_course.year
                elif first_course.year > second_course.year:
                    second_year = second_course.year
                else:
                    first_year = first_course.year
                    second_year = second_course.year
            except AttributeError:
                pass

            course_fall_term = datetime.datetime(year=first_year, month=8, day=24)
            course_spring_term = datetime.datetime(year=first_year, month=1, day=11)
            summer_1st_term = datetime.datetime(year=second_year, month=5, day=15)
            summer_2nd_term = datetime.datetime(year=second_year, month=6, day=18)

            if (first_course is "None") or (first_course is None) or (second_course is "None") or \
                    (second_course is None):
                return False
            else:
                # Checks if there is a difference bigger than 33 days in the dates.
                if ((course_fall_term - datetime.timedelta(days=33)) <= first_course <=
                    (course_fall_term + datetime.timedelta(days=33))
                    or (course_spring_term - datetime.timedelta(days=33)) <= first_course
                    <= (course_spring_term + datetime.timedelta(days=33))) \
                        and ((course_fall_term - datetime.timedelta(days=33)) <= second_course
                             <= (course_fall_term + datetime.timedelta(days=33))
                             or (course_spring_term - datetime.timedelta(days=33)) <= second_course
                             <= (course_spring_term + datetime.timedelta(days=33))):
                    return False
                elif first_course.month == second_course.month:
                    return False
                # Summer
                elif ((summer_1st_term - datetime.timedelta(days=33)).month <= first_course.month <=
                     (summer_1st_term + datetime.timedelta(days=33)).month
                     or (summer_2nd_term - datetime.timedelta(days=33)).month <= first_course.month
                     <= (summer_2nd_term + datetime.timedelta(days=33)).month) \
                    and ((summer_1st_term - datetime.timedelta(days=33)).month <= second_course.month
                         <= (summer_1st_term + datetime.timedelta(days=33)).month
                         or (summer_2nd_term - datetime.timedelta(days=33)).month <= second_course.month
                         <= (summer_2nd_term + datetime.timedelta(days=33)).month):
                    return False
                else:
                    return True

        def check_room_capacity(course, dict_room_cap):
            for rooms, rooms_cap in dict_room_cap.items():
                if (rooms is not None) & (course.get("Room") is not None):
                    if course.get("Room") == rooms:
                        if course.get("Enrollment") == rooms_cap:
                            pass
                        else:
                            try:
                                if int(course.get("Enrollment")) > int(rooms_cap):
                                    return course, rooms_cap, "FEBBBB"
                                else:
                                    return course, rooms_cap, "C5C5FF"
                            except ValueError:
                                pass

        # Creates a copy of our main dict
        list_dict_copy = self.dict_courses_list.copy()
        for course_i in range(len(list_dict_copy)):

            try:
                courses, room_cap, color = check_room_capacity(list_dict_copy[course_i], self.classroom_capacity)
                if (courses is not None) & (room_cap is not None):
                    comment = courses.get("Room") + " capacity expected to be " + \
                              str(room_cap) + " for " + courses.get("Course") + (" " * 140)
                    self.create_report_dictionary(courses.get("Row"), 9, courses.get("File"), courses.get("Sheet_Name"),
                                                  color, comment)
            except TypeError:
                pass
            for course_d in range(len(list_dict_copy) - 1):
                if course_i != (course_d + 1):
                    try:
                        # Checking if the courses are hybrid
                        if list_dict_copy[course_i].get("Course") == list_dict_copy[course_d + 1].get("Course"):
                            if list_dict_copy[course_i].get("Course_Days"):
                                if not list_dict_copy[course_d + 1].get("Course_Days"):
                                    list_dict_copy[course_i].get("Type").append("Hybrid")
                                    # Changing previous Online type to Hybrid
                                    list_dict_copy[course_d + 1]["Type"] = "Hybrid"

                        # Checks Course days similarities
                        if not list(set(list_dict_copy[course_i].get("Course_Days")) &
                                    set(list_dict_copy[course_d + 1].get("Course_Days"))):
                            pass
                        else:
                            start_time_i = list_dict_copy[course_i].get('Start_Time')
                            if start_time_i.upper() == "ONLINE":
                                start_time_i = None
                            start_time_d = list_dict_copy[course_d + 1].get('Start_Time')
                            if start_time_d.upper() == "ONLINE":
                                start_time_d = None
                            end_time_i = list_dict_copy[course_i].get('End_Time')
                            end_time_d = list_dict_copy[course_d + 1].get('End_Time')

                            start_date_i = list_dict_copy[course_i].get('Start_Date')
                            start_date_d = list_dict_copy[course_d + 1].get("Start_Date")

                            if start_time_i is None or start_time_d is None:
                                pass
                            elif start_time_i.upper() == 'NONE' or start_time_d.upper() == 'NONE':
                                pass
                            elif end_time_i is None or end_time_d is None:
                                pass
                            elif end_time_i.upper() == 'NONE' or end_time_d.upper() == 'NONE':
                                pass
                            else:
                                if (list_dict_copy[course_i].get('Room') is not None) & \
                                        (list_dict_copy[course_d + 1].get('Room') is not None):
                                    room_ig = "".join(list_dict_copy[course_i].get('Room').split())
                                    room_d = "".join(list_dict_copy[course_d + 1].get('Room').split())

                                    section_numb_ig = list_dict_copy[course_i].get('Course').split("-")
                                    section_numb_d = list_dict_copy[course_d + 1].get("Course").split("-")

                                    # It is normal for section 40 or 41 to conflict with another course
                                    if (section_numb_d[1] == '40') or (section_numb_ig[1] == '40'):
                                        pass
                                    elif (section_numb_d[1] == '41') or (section_numb_ig[1] == '41'):
                                        pass

                                    elif room_ig == room_d:
                                        # Checks for dates
                                        if check_course_dates(start_date_i, start_date_d) is True:
                                            self.different_date_list.append(self.dict_courses_list[course_d + 1])
                                            del self.dict_courses_list[course_d + 1]
                                        else:
                                            if check_course_dates(start_date_i, start_date_d) is False:
                                                # Transforms variables to float
                                                start_time_i = (float(start_time_i[0:2] + '.' + start_time_i[3:5]))
                                                start_time_d = (float(start_time_d[0:2] + '.' + start_time_d[3:5]))
                                                end_time_i = (float(end_time_i[0:2] + '.' + end_time_i[3:5]))
                                                end_time_d = (float(end_time_d[0:2] + '.' + end_time_d[3:5]))

                                                # Checks for conflicts
                                                if start_time_d <= start_time_i <= end_time_d:
                                                    self.time_conflict_comment(list_dict_copy[course_i],
                                                                               list_dict_copy[course_d + 1])

                                                elif start_time_d <= end_time_i <= end_time_d:
                                                    self.time_conflict_comment(list_dict_copy[course_i],
                                                                               list_dict_copy[course_d + 1])
                                                    # del self.list_dict_courses[course_i]
                                                else:
                                                    # Checks if the courses have less than 15 min difference if so
                                                    # mark as conflict
                                                    fifteenth_minutes_start_i = start_time_i
                                                    fifteenth_minutes_end_i = end_time_i
                                                    fifteenth_minutes_start_d = start_time_d - 0.14
                                                    fifteenth_minutes_end_d = end_time_d + 0.14

                                                    def check_minutes(time):
                                                        # transforms minutes to hours if over 60
                                                        if math.modf(time)[0] >= 0.60:
                                                            x = math.modf(time)[1] + 1 + math.modf(time)[0] - 0.60
                                                            return x
                                                        else:
                                                            return time

                                                    fifteenth_minutes_start_i = check_minutes(
                                                        fifteenth_minutes_start_i)
                                                    fifteenth_minutes_end_i = check_minutes(
                                                        fifteenth_minutes_end_i)
                                                    fifteenth_minutes_start_d = check_minutes(
                                                        fifteenth_minutes_start_d)
                                                    fifteenth_minutes_end_d = check_minutes(fifteenth_minutes_end_d)
                                                    if fifteenth_minutes_start_d <= fifteenth_minutes_start_i <= \
                                                            fifteenth_minutes_end_d:
                                                        self.time_conflict_comment(list_dict_copy[course_i],
                                                                                   list_dict_copy[course_d + 1], True)
                                                    elif fifteenth_minutes_start_d <= fifteenth_minutes_end_i <= \
                                                            fifteenth_minutes_end_d:
                                                        self.time_conflict_comment(list_dict_copy[course_i],
                                                                                   list_dict_copy[course_d + 1], True)
                                                    else:
                                                        pass
                    except IndexError:
                        pass

    def time_conflict_comment(self, first_dict, second_dict, fifteenth_conflict=False):
        """Creates a dictionary if program finds a conflict"""
        first_dict_comment = first_dict.get("Course") + ": "
        second_dict_comment = second_dict.get("Course") + ": "

        # Depending on your conflict assigns different colors and comments
        if fifteenth_conflict is True:
            color = "FEBBBB"
            first_dict_comment += "time difference is less than 15 min with " + second_dict.get("Course") + (" " * 130)
            second_dict_comment += "time difference is less than 15 min with " + first_dict.get("Course") + (" " * 130)
        else:
            color = "FF687B"
            first_dict_comment += "conflicts with " + second_dict.get("Course") + (" " * 150)
            second_dict_comment += "conflicts with " + first_dict.get("Course") + (" " * 150)

        self.create_report_dictionary(
            excel_row=first_dict.get("Row"),
            excel_column=1,
            excel_file_path=first_dict.get("File"),
            excel_sheet_name=first_dict.get("Sheet_Name"),
            error_color=color, error_comment=first_dict_comment)

        self.create_report_dictionary(
            excel_row=second_dict.get("Row"),
            excel_column=1,
            excel_file_path=second_dict.get("File"),
            excel_sheet_name=second_dict.get("Sheet_Name"),
            error_color=color, error_comment=second_dict_comment)

    def create_excel_table(self):
        """Moves into another class"""
        room_schedule_table.MasterDesign(self.save_path, self.dict_courses_list, self.different_date_list,
                                         self.course_days, self.table_year, self.table_name, self.table_semester)

    def get_excel_errors(self):
        """Returns founded errors"""
        return self.user_excel_errors
