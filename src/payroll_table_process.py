from nameparser import HumanName
import requests
from bs4 import BeautifulSoup
import re
import datetime
import os
import csv
import openpyxl
import string
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Font, Alignment, PatternFill


class PayrollTable:
    def __init__(self, c_dict, folder_dir):
        self.save_path = folder_dir
        self.courses_dict = c_dict

        self.allowed_departments = [
            'Accounting', 'Business Law', 'Business', 'Finance',
            'Marketing & International Business', 'MBA', 'MACC',
            'Management', 'ACCT.BLAW.MACC', 'MRKT.IBUS', 'MGMT.MBA'
        ]

        self.professor_data = []
        self.date_issue_list = []
        # self.breaking_pages = []
        self.faculty_dict = dict()
        self.payroll_dict_format()
        self.create_excel_table()

    def payroll_dict_format(self):
        def create_faculty_dict(courses_dict):
            result = dict()
            for course_len in range(len(courses_dict)):
                for dict_key, faculty_name in courses_dict[course_len].items():
                    if dict_key == "Faculty":
                        if faculty_name != "None":
                            result.setdefault(faculty_name, [])
                            result[faculty_name].append(courses_dict[course_len])
            return result

        def get_professor_type(f_dict):
            prof_data = []
            allowed_departments = [
                'Accounting', 'Business Law', 'Business', 'Finance',
                'International Business', 'MBA', 'MACC',
                'Management', 'Marketing'
            ]

            def fetch_page():
                url = f'http://www.mnsu.edu/find/people.php?givenname={first_name}&sn={last_name}&employeetype='
                response = requests.get(url)
                return response.text

            def fetch_professor_type():
                page = fetch_page()
                soup = BeautifulSoup(page, 'html.parser')
                for p in soup.find_all('p'):
                    match = re.search(
                        fr'Name:(?P<name>.+?)Department:(?P<department>.+?{".*?|.*?".join(allowed_departments)}.+?); '
                        fr'Title: (?P<title>.+?); Type: (?P<type>.+?)Address: (?P<address>.+?); '
                        fr'Phone: (?P<phone>.+?); Email: (?P<email>.+?);',
                        p.text, flags=re.I
                    )

                    if match:
                        # let's remove trailing whitespaces in extracted data
                        data = match.groupdict()
                        for key, value in data.items():
                            data[key] = value.strip()

                        return data

                return None

            for professor in f_dict:
                name = HumanName(professor)
                first_name = name.first
                last_name = name.last
                professor_type = fetch_professor_type()
                if professor_type is None:
                    first_name = ""
                    professor_type = fetch_professor_type()
                    if professor_type is not None:
                        prof_data.append(professor_type)
                else:
                    prof_data.append(professor_type)
            return prof_data

        faculty_dict = create_faculty_dict(self.courses_dict)
        self.professor_data = get_professor_type(faculty_dict)

        def get_csv_file(file):
            cost_center = dict()
            if os.path.isfile(file):
                with open(file) as csv_file:
                    read_csv_file = csv.DictReader(csv_file, delimiter=',')
                    for row in read_csv_file:
                        cost_center = dict(row)
            return cost_center

        csv_file_data = get_csv_file('department_cost.csv')
        final_dict = dict()
        for professor, courses in faculty_dict.items():
            first_name = HumanName(professor).first
            last_name = HumanName(professor).last
            current_professor = next(
                (i for i in self.professor_data
                 if re.search(fr'{first_name}.*?{last_name}', i['name'])),
                dict()
            )
            final_dict[professor] = dict()
            final_dict[professor]['professor'] = current_professor
            tmp_courses = []

            for course in courses:
                if course['Department'].lower() == 'business':
                    cost = csv_file_data.get(current_professor.get('department'))
                    if current_professor.get('department'):
                        course['Department'] = current_professor['department']
                else:
                    cost = csv_file_data.get(course['Department'])

                # TODO: ADD BUTTON IF USER WANT TO SEPERATE ALL DEPARTMENTS
                if (course['Department'].lower() == 'accounting') or \
                        (course['Department'].lower() == 'business law') or \
                        (course['Department'].lower() == 'macc'):
                    course['Department'] = "ACCT.BLAW.MACC"
                elif (course['Department'].lower() == 'marketing') or \
                        (course['Department'].lower() == 'international business') or \
                        (course['Department'].lower() == 'marketing & international business'):
                    course['Department'] = "MRKT.IBUS"
                elif (course['Department'].lower() == 'management') or \
                        (course['Department'].lower() == 'mba'):
                    course['Department'] = "MGMT.MBA"
                else:
                    pass

                course['Cost'] = cost
                tmp_courses.append(course)

                # Semester change
                try:
                    course_year = course.get("Start_Date").year
                    fall_start_date = datetime.datetime(year=course_year, month=8, day=26)
                    spring_start_date = datetime.datetime(year=course_year, month=1, day=13)
                    summer1_start_date = datetime.datetime(year=course_year, month=5, day=18)
                    summer2_start_date = datetime.datetime(year=course_year, month=6, day=22)

                    if (fall_start_date - datetime.timedelta(days=33)) <= (course.get("Start_Date")) <= \
                            (fall_start_date + datetime.timedelta(days=83)):
                        course["Semester"] = "Fall"
                        course["Year"] = course_year

                    elif (spring_start_date - datetime.timedelta(days=33)) <= (course.get("Start_Date")) <= \
                            (spring_start_date + datetime.timedelta(days=83)):
                        course["Semester"] = "Spring"
                        course["Year"] = course_year

                    elif (summer1_start_date - datetime.timedelta(days=21)) <= (course.get("Start_Date")) <= \
                            (summer1_start_date + datetime.timedelta(days=21)):
                        course["Semester"] = "Summer1"
                        course["Year"] = course_year

                    elif (summer2_start_date - datetime.timedelta(days=21)) <= (course.get("Start_Date")) <= \
                            (summer2_start_date + datetime.timedelta(days=21)):
                        course["Semester"] = "Summer2"
                        course["Year"] = course_year

                    else:
                        course["Semester"] = None
                        course["Year"] = course_year
                except AttributeError:
                    self.date_issue_list.append(course)

            final_dict[professor]['courses'] = tmp_courses

        self.faculty_dict = final_dict

    def print_department_info(self, department):
        filtered_dict_courses = dict()

        for professor, data in self.faculty_dict.items():
            filtered_courses = [course for course in data['courses'] if
                                re.search(fr'.*?{department}.*?', course['Department'])]
            if filtered_courses:
                filtered_dict_courses[professor] = dict()
                filtered_dict_courses[professor]['courses'] = filtered_courses
                filtered_dict_courses[professor]['professor'] = data.get("professor")

        return filtered_dict_courses

    def create_excel_table(self):
        for dep in self.allowed_departments:
            if dep.lower() not in ["accounting", "business law", "marketing", "international business", "macc",
                                   "management", "mba"]:
                department_courses = self.print_department_info(dep)
                if department_courses:
                    self.main_program_controller(dep, department_courses)

    def main_program_controller(self, department, department_courses):
        self.dep_courses = department_courses
        self.create_excel_file()

        self.create_excel_sheet("Payroll Table", True)
        self.dates_heading()
        self.instructions_text()
        self.department_text(department)
        self.merge_headings()
        self.insert_courses()
        self.adjust_cells_width()
        self.border_all_cells()

        self.workbook.save(self.save_path + "\\" + department + "_" + "Payroll" + ".xlsx")

    def create_excel_file(self):
        def create_directory():
            """Creates directory for created excel file"""
            if not os.path.exists('__excel_files'):
                os.makedirs('__excel_files')

        create_directory()

        self.workbook = openpyxl.Workbook()

    def create_excel_sheet(self, sheet_name, first_sheet=False):
        if first_sheet is True:
            self.sheet = self.workbook["Sheet"]
            self.sheet.title = "Payroll Table"
        else:
            self.workbook.create_sheet(sheet_name)
            self.sheet = self.workbook.get_sheet_by_name(sheet_name)

    def dates_heading(self):
        self.sheet["A1"] = "Table created:"
        self.sheet["B1"] = datetime.datetime.today().strftime('%m/%d/%Y')

    def instructions_text(self):
        self.sheet["G1"] = "* Key Concept for Scheduling - *Max Credits for Fall=14"
        self.sheet["G1"].font = Font(color="FF493B")

        self.sheet["G2"] = "* Final Total Credits must be 24"
        self.sheet["G2"].font = Font(color="FF493B")

        self.sheet["G3"] = "* Max Credit for the Year 29 with Overload"
        self.sheet["G3"].font = Font(color="FF493B")

    def department_text(self, department):
        if department == "ACCT.BLAW.MACC":
            self.sheet["A3"] = "Accounting, Business Law and MACC"

        elif department == "MRKT.IBUS":
            self.sheet["A3"] = "Marketing and International Business"

        elif department == "MGMT.MBA":
            self.sheet["A3"] = "Managements and MBA"

        else:
            self.sheet["A3"] = department

    def merge_headings(self):
        # Date
        self.merge_excel_cells(start_row=1, start_column=2, end_row=1, end_column=2)

        # Department
        self.merge_excel_cells(start_row=3, start_column=1, end_row=3, end_column=3)

        # Instructions merge cells
        self.merge_excel_cells(start_row=1, start_column=7, end_row=1, end_column=11)
        self.merge_excel_cells(start_row=2, start_column=7, end_row=2, end_column=11)
        self.merge_excel_cells(start_row=3, start_column=7, end_row=3, end_column=11)

    def insert_courses(self):
        changing_row = 4

        # Columns

        def insert_columns(semester, year, sheet, row):
            repetitive_headers = ["Subject", "Section", "Course Name", "Cost Center", "Credits", "Enrollments", "Room",
                                  "Time", "Days", "Dates"]
            sheet["D" + str(row)] = semester.upper() + " " + str(year)
            sheet["D" + str(row)].font = Font(sz=13, bold=True, italic=False)
            sheet["D" + str(row)].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            row += 1

            for i in range(len(repetitive_headers)):
                column_letter = ''.join(string.ascii_uppercase[i + 1])
                sheet[column_letter + str(row)] = repetitive_headers[i]
                sheet[column_letter + str(row)].font = Font(sz=11, bold=True, italic=False)
                sheet[column_letter + str(row)].alignment = Alignment(horizontal='center', vertical='center',
                                                                      wrap_text=True)
            row += 1
            return row

        def remove_dict_duplicates(dict_list):
            new_dict_list = []
            # removing keys
            for i in range(len(dict_list)):
                dict_list[i].pop('Row', None)
                dict_list[i].pop('File', None)
                dict_list[i].pop('Sheet_Name', None)

            for i in range(len(dict_list)):
                if dict_list[i] not in dict_list[i + 1:]:
                    new_dict_list.append(dict_list[i])

            return new_dict_list

        def insert_dict(sheet, course, row):
            # 10 columns
            c = course.get("Course").split()
            sheet["B" + str(row)] = c[0]
            sheet["B" + str(row)].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            sheet["C" + str(row)] = c[1]
            sheet["C" + str(row)].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            sheet["D" + str(row)] = course.get("Course_Title")
            sheet["D" + str(row)].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            try:
                sheet["E" + str(row)] = int(course.get("Cost"))
                sheet["E" + str(row)].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

                sheet["F" + str(row)] = int(course.get("Credits"))
                sheet["F" + str(row)].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            except (ValueError, TypeError):
                sheet["E" + str(row)] = course.get("Cost")
                sheet["E" + str(row)].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

                sheet["F" + str(row)] = course.get("Credits")
                sheet["F" + str(row)].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            sheet["G" + str(row)] = course.get("Enrollment")
            sheet["G" + str(row)].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            sheet["H" + str(row)] = course.get("Room")
            sheet["H" + str(row)].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            sheet["I" + str(row)] = course.get("Start_Time") + " - " + course.get("End_Time")
            sheet["I" + str(row)].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            courses_days_string = ", ".join(course.get("Course_Days"))
            sheet["J" + str(row)] = courses_days_string
            sheet["J" + str(row)].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            try:
                sheet["K" + str(row)] = course.get("Start_Date").strftime('%m/%d/%Y') + " - " + course.get(
                    "End_Date").strftime('%m/%d/%Y')
                sheet["K" + str(row)].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            except AttributeError:
                pass

            row += 1
            return row

        adjunct_courses = []

        unassigned_prof_list = []

        fall_year = None
        spring_year = None
        # Employee
        for faculty in self.dep_courses:
            s_row = changing_row
            fall_course_list = []
            spring_course_list = []
            if (faculty == '\xa0') or (faculty[0:10].lower() == "unassigned") or (faculty == "") or \
                    (faculty == " ") or (faculty == "None") or (faculty is None) or (faculty[0:5].lower() == "staff"):
                for course in (self.dep_courses.get(faculty)).get("courses"):
                    unassigned_prof_list.append(course)
                break
            employee_type = ((self.dep_courses.get(faculty)).get("professor")).get("title")
            if (employee_type is None) or (employee_type.lower() != "adjunct"):
                self.sheet["A" + str(changing_row)] = re.sub(' +', ' ', faculty)  # Removes spaces
                for course in (self.dep_courses.get(faculty)).get("courses"):
                    if course.get("Semester") == "Fall":
                        fall_course_list.append(course)
                        fall_year = course.get("Year")
                    elif course.get("Semester") == "Spring":
                        spring_course_list.append(course)
                        spring_year = course.get("Year")
                    else:
                        self.date_issue_list.append(course)

                if (fall_year is None) or (spring_year is None):
                    if fall_year is None:
                        now = datetime.datetime.now()
                        fall_year = (now.year) - 1
                    else:
                        now = datetime.datetime.now()
                        spring_year = now.year

                # Remove duplicates
                fall_course_list = remove_dict_duplicates(fall_course_list)
                spring_course_list = remove_dict_duplicates(spring_course_list)

                changing_row = insert_columns("FALL", fall_year, self.sheet, changing_row)
                f_credits = 0
                for f_courses in fall_course_list:
                    f_credits += f_courses.get("Credits")
                    changing_row = insert_dict(self.sheet, f_courses, changing_row)
                self.sheet["F" + str(changing_row)] = f_credits
                clr = PatternFill(start_color='fcdfbe', end_color='fcdfbe', fill_type='solid')
                self.sheet["F" + str(changing_row)].fill = clr
                self.sheet["F" + str(changing_row)].alignment = Alignment(horizontal='center', vertical='center',
                                                                          wrap_text=True)
                changing_row += 1
                changing_row = insert_columns("SPRING", spring_year, self.sheet, changing_row)

                s_credits = 0
                for c_courses in spring_course_list:
                    s_credits += c_courses.get("Credits")
                    changing_row = insert_dict(self.sheet, c_courses, changing_row)
                self.sheet["F" + str(changing_row)] = s_credits
                clr = PatternFill(start_color='fcdfbe', end_color='fcdfbe', fill_type='solid')
                self.sheet["F" + str(changing_row)].fill = clr
                self.sheet["F" + str(changing_row)].alignment = Alignment(horizontal='center', vertical='center',
                                                                          wrap_text=True)
                changing_row += 1

                # Total credits for both semesters
                total_credits = f_credits + s_credits
                self.sheet["F" + str(changing_row)] = total_credits
                if total_credits > 24:
                    clr = PatternFill(start_color='fa8072', end_color='fa8072', fill_type='solid')
                else:
                    clr = PatternFill(start_color='76C54F', end_color='d5d8e0', fill_type='solid')
                self.sheet["F" + str(changing_row)].fill = clr
                self.sheet["F" + str(changing_row)].font = Font(sz=11, bold=True, italic=False)
                self.sheet["F" + str(changing_row)].alignment = Alignment(horizontal='center', vertical='center',
                                                                          wrap_text=True)

                changing_row += 1

                self.merge_excel_cells(start_row=s_row, start_column=1, end_row=changing_row - 1, end_column=1,
                                       style=True, bold=True)
            else:
                adjunct_courses.append(self.dep_courses[faculty])
        # Adjunct Section
        for adj_courses in adjunct_courses:
            fall_course_list = []
            spring_course_list = []
            for course in (adj_courses.get("courses")):
                if course.get("Semester") == "Fall":
                    fall_course_list.append(course)
                elif course.get("Semester") == "Spring":
                    spring_course_list.append(course)
                else:
                    self.date_issue_list.append(course)

            # Remove duplicates
            fall_course_list = remove_dict_duplicates(fall_course_list)
            spring_course_list = remove_dict_duplicates(spring_course_list)

            # Adjunct heading
            self.sheet["A" + str(changing_row)] = "ADJUNCT"
            self.sheet["A" + str(changing_row)].font = Font(sz=13, bold=True, italic=False)
            clr = PatternFill(start_color='ffbc25', end_color='ffbc25', fill_type='solid')
            self.sheet["A" + str(changing_row)].fill = clr
            changing_row += 1

            # Semester headings
            if fall_course_list:
                self.sheet["A" + str(changing_row)] = "FALL " + str(fall_year)
                self.sheet["A" + str(changing_row)].font = Font(sz=11, bold=True, italic=False)
                clr = PatternFill(start_color='b7dbdd', end_color='b7dbdd', fill_type='solid')
                self.sheet["A" + str(changing_row)].fill = clr
                changing_row += 1
                for f_course in fall_course_list:
                    self.sheet["A" + str(changing_row)] = f_course.get("Faculty")
                    self.sheet["A" + str(changing_row)].alignment = Alignment(horizontal='center', vertical='center',
                                                                              wrap_text=True)
                    changing_row = insert_dict(self.sheet, f_course, changing_row)

            if spring_course_list:
                self.sheet["A" + str(changing_row)] = "SPRING " + str(spring_year)
                self.sheet["A" + str(changing_row)].font = Font(sz=11, bold=True, italic=False)
                clr = PatternFill(start_color='d1dccf', end_color='d1dccf', fill_type='solid')
                self.sheet["A" + str(changing_row)].fill = clr
                changing_row += 1
                for s_course in spring_course_list:
                    self.sheet["A" + str(changing_row)] = s_course.get("Faculty")
                    self.sheet["A" + str(changing_row)].alignment = Alignment(horizontal='center', vertical='center',
                                                                              wrap_text=True)
                    changing_row = insert_dict(self.sheet, s_course, changing_row)

        # Error section
        if unassigned_prof_list:
            # Remove duplicates
            unassigned_prof_list = remove_dict_duplicates(unassigned_prof_list)
            changing_row += 1
            self.sheet["A" + str(changing_row)] = "Unassigned Professors Courses:"
            clr = PatternFill(start_color='ff493b', end_color='ff493b', fill_type='solid')
            self.sheet["A" + str(changing_row)].fill = clr
            self.merge_excel_cells(changing_row, 1, changing_row, 3, True)
            changing_row += 1
            for u_prof in unassigned_prof_list:
                self.sheet["A" + str(changing_row)] = "Not found"
                changing_row = insert_dict(self.sheet, u_prof, changing_row)

        # Date issue courses
        if self.date_issue_list:
            changing_row += 1
            self.sheet["A" + str(changing_row)] = "Couldn't insert following courses:"
            clr = PatternFill(start_color='ff493b', end_color='ff493b', fill_type='solid')
            self.sheet["A" + str(changing_row)].fill = clr
            self.merge_excel_cells(changing_row, 1, changing_row, 3, True)
            changing_row += 1
            for d_prof in self.date_issue_list:
                self.sheet["A" + str(changing_row)] = d_prof.get("Faculty")
                changing_row = insert_dict(self.sheet, d_prof, changing_row)

    def adjust_cells_width(self):
        """Adjusts all the cell width. It is really cool"""
        # Gets last column
        for column in self.sheet.columns:
            max_length = 0
            # Gets column coordinates
            get_column = column[0].column
            if get_column is "A":
                self.sheet.column_dimensions["A"].width = 18
            elif get_column is "G":
                self.sheet.column_dimensions["G"].width = 13
            else:
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except TypeError:
                        pass
                # A formula for auto adjusted width
                adjusted_width = (max_length + 2) * 1.035
                if adjusted_width > 25.5:
                    adjusted_width = 25.5
                self.sheet.column_dimensions[get_column].width = adjusted_width

    def merge_excel_cells(self, start_row, start_column, end_row, end_column, style=False, bold=False):
        """Merges excel cells"""
        excel_sheet = self.sheet
        excel_sheet.merge_cells(start_row=start_row, start_column=start_column, end_row=end_row, end_column=end_column)

        def style_excel_cell(sheet, row, column):
            """Styles a cell"""
            sheet.cell(row=row, column=column).font = Font(sz=11, bold=bold, italic=False)
            sheet.cell(row=row, column=column).alignment = Alignment(horizontal='center',
                                                                     vertical='center', wrap_text=True)

        if style is True:
            style_excel_cell(excel_sheet, start_row, start_column)

    def border_all_cells(self):
        """Borders all table"""
        # Gets table size
        excel_max_row = self.sheet.max_row
        excel_max_column = self.sheet.max_column

        # Style of a border
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))

        # Goes over each cell and applies border
        for column in range(excel_max_column):
            # Transfers column to an alphabetical format
            col_letter = ''.join(string.ascii_uppercase[column])
            for row in range(excel_max_row):
                row += 1
                if row in [1, 2, 3]:
                    pass
                else:
                    self.sheet[col_letter + str(row)].border = thin_border
