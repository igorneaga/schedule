import os
import re
import string

import openpyxl
from openpyxl.comments import Comment
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.styles.borders import Border, Side
from openpyxl.worksheet.pagebreak import Break


class MasterDesign:

    def __init__(self, list_dict_courses, days, year, table_name, table_semester):
        self.list_dict_courses = list_dict_courses
        self.days = days
        self.table_year = year
        self.table_name = table_name
        self.table_semester = table_semester

        self.workbook = None
        self.sheet = None

        # College of Business rooms
        self.cob_rooms = ["MH 0102", "MH 0208", "MH 0209", "MH 0210", "MH 0211", "AH 0205", "AH 0209", "AH 0216", "AH 0220", "AH 0320"]

        self.list_unique_times = []
        self.course_types_list = []

        self.main_program_controller()

    def main_program_controller(self):
        self.create_master_file()
        # Classroom table section
        self.create_excel_sheet(sheet_name="Classroom Table", first_sheet=True)
        self.set_time_row()
        self.set_courses()

        self.color_cell_meaning()

        self.set_excel_heading(heading_name="Classroom Table")
        self.adjust_cells_width()
        self.set_page_break()

        # Online table section
        # self.create_excel_sheet(sheet_name="Online Table")
        # self.set_excel_heading(heading_name="Online Table")
        # Hybrid table section
        # self.create_excel_sheet(sheet_name="Hybrid Table")
        # self.set_excel_heading(heading_name="Hybrid Table")
        # MBA MACC table section
        # self.create_excel_sheet(sheet_name="MBA MACC Table")
        # self.set_excel_heading(heading_name="MBA MACC Table")
        # Telepresence table section
        # self.create_excel_sheet(sheet_name="Telepresence Table")
        # self.set_excel_heading(heading_name="Telepresence Table")

        # Not Included Courses table section
        # self.create_excel_sheet(sheet_name="Not Included Courses")
        # self.set_excel_heading(heading_name="Not Included Courses")

        # Note: Save only once
        self.save_excel_file()

    def save_excel_file(self):
        """Saves excel file by using user input"""
        if self.table_name[-5:] == ".xlsx":
            self.workbook.save('__excel_files\\' + self.table_name)
        else:
            self.table_name = "".join(self.table_name.split())
            if not self.table_name:
                self.table_name = "Empty_Name"
            self.workbook.save('__excel_files\\' + self.table_name + ".xlsx")

    def create_master_file(self):
        """Creates a folder and excel file"""
        def create_directory():
            """Creates directory for created excel file"""
            if not os.path.exists('__excel_files'):
                os.makedirs('__excel_files')
        create_directory()

        self.workbook = openpyxl.Workbook()

    def create_excel_sheet(self, sheet_name, first_sheet=False):
        """Creates or renames the sheet name"""
        if first_sheet is True:
            self.sheet = self.workbook.get_sheet_by_name('Sheet')
            self.sheet.title = sheet_name
        else:
            self.workbook.create_sheet(sheet_name)
            self.sheet = self.workbook.get_sheet_by_name(sheet_name)

    def set_excel_heading(self, heading_name):
        """Sets excel heading based on a user input"""

        self.sheet.oddHeader.center.text = str(heading_name) + " of " + str(self.table_semester) + " " + str(self.
                                                                                                             table_year)
        self.sheet.oddHeader.center.size = 14

        def set_course_term(semester, year, sheet):
            """Sets a course semester and year on a first cell"""
            sheet.merge_cells("A1:B1")
            sheet["A1"] = "Term: " + str(semester) + " " + str(year)
            sheet["A1"].font = Font(sz=11, bold=True, italic=False)
            sheet["A1"].alignment = Alignment(horizontal='center', vertical='center')
        set_course_term(self.table_semester, self.table_year, self.sheet)

    def prepare_row_time(self):
        """Prepares time to have a valid order and valid value. Needs only for classroom table"""

        list_times = []

        # Takes out of Online courses
        for i in range(len(self.list_dict_courses)):
            if self.list_dict_courses[i].get("Start_Time") != "Online":
                list_times.append(self.list_dict_courses[i].get("Start_Time"))
            if self.list_dict_courses[i].get("End_Time") != "Online":
                list_times.append(self.list_dict_courses[i].get("End_Time"))

        # Takes out duplicates
        list_times = list(set(list_times))

        def set_time_order(time_list):
            """Making time be on the correct order."""
            morning_time = []
            evening_time = []
            for t in range(len(time_list)):
                # Earliest class starts at 8
                if any(c in time_list[t][0:2] for c in ("08", "09", "10", "11", "12")):
                    morning_time.append(time_list[t][0:5])
                # The latest class can start at 6 or 7
                if any(c in time_list[t][0:2] for c in ("01", "02", "03", "04", "05", "06", "07")):
                    evening_time.append(time_list[t][0:5])

            morning_time.sort()
            evening_time.sort()
            # Combines them
            row_time = morning_time + evening_time
            return row_time

        list_times = set_time_order(list_times)
        return list_times

    def set_time_row(self):
        """Creates a row of unique times. Needs only for classroom table"""
        list_unique_times = self.prepare_row_time()
        temp_time_dict = dict()
        time_row_column = 2
        for t in range(len(list_unique_times)):
            alphabet = ''.join(string.ascii_uppercase[time_row_column])
            time_row = str(alphabet) + '1'
            self.sheet[time_row] = list_unique_times[t]
            self.sheet[time_row].font = Font(sz=11, bold=True, italic=False)
            self.sheet[time_row].alignment = Alignment(horizontal='center', vertical='center')
            temp_time_dict["Time"] = list_unique_times[t]
            temp_time_dict["Column_Num"] = time_row_column
            self.list_unique_times.append(temp_time_dict.copy())
            time_row_column += 1

    def adjust_cells_width(self):
        """Adjusts all the cell width. It is really cool"""
        # Gets last column
        excel_max_column = self.sheet.max_column

        # Transfers column to an alphabetical format
        col_letter = ''.join(string.ascii_uppercase[excel_max_column - 4])

        for column in self.sheet.columns:
            max_length = 0
            # Gets column coordinates
            get_column = column[0].column
            # Column "A" and "B" will have a standard size
            if get_column is "A":
                self.sheet.column_dimensions["A"].width = 12
            elif get_column is "B":
                self.sheet.column_dimensions["B"].width = 12
            else:
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except TypeError:
                        pass
                # A formula for auto adjusted width
                adjusted_width = (max_length + 2) * 1.05
                # Limits adjusted width
                if get_column is col_letter:
                    # Last column needs to be bigger to fit everything correctly
                    if adjusted_width > 24:
                        adjusted_width = 24
                elif adjusted_width > 14:
                    adjusted_width = 14

                self.sheet.column_dimensions[get_column].width = adjusted_width

    def merge_excel_cells(self, start_row, start_column, end_row, end_column, style=False, bold=False):
        """Merges excel cells"""
        excel_sheet = self.sheet
        excel_sheet.merge_cells(start_row=start_row, start_column=start_column, end_row=end_row, end_column=end_column)

        def style_excel_cell(sheet, row, column):
            """Styles a cell"""
            sheet.cell(row=row, column=column).font = Font(sz=11, bold=bold, italic=False)
            sheet.cell(row=row, column=column).alignment = Alignment(horizontal='center',
                                                                     vertical='center', wrap_text=True)
        if style is True:
            style_excel_cell(excel_sheet, start_row, start_column)

    def set_courses(self):
        """Sets courses in a excel file"""

        # An excel row number where courses can be placed
        start_excel_row = 2

        def set_room_dict(list_courses):
            """Creates a new dictionary where the key is a room number and contains all courses related to it"""
            result = dict()
            for course_len in range(len(list_courses)):
                for k, course_value in list_courses[course_len].items():
                    if k == "Room":
                        result.setdefault(course_value, [])
                        result[course_value].append(list_courses[course_len])
            return result

        def set_room_number(sheet, room_row, room):
            """Sets a room in a "A" column"""
            sheet['A' + str(room_row)] = room

        def set_days_row(sheet, day_row, days, b_num):
            """Sets days in a "B" column """
            sheet['B' + str(day_row)] = days[b_num]

        def get_cell_value(get_column, sheet, get_row):
            """Returns a cell value"""
            return sheet[get_column + str(get_row)].value

        def border_all_cells(sheet):
            """Borders all table"""
            # Gets table size
            excel_max_row = sheet.max_row
            excel_max_column = sheet.max_column

            # Transfers column to an alphabetical format
            col_letter = ''.join(string.ascii_uppercase[excel_max_column-1])

            # Gets full coordinates of a table
            full_cord = "A1:" + str(col_letter) + str(excel_max_row)

            # Style of a border
            thin_border = Border(left=Side(style='thin'),
                                 right=Side(style='thin'),
                                 top=Side(style='thin'),
                                 bottom=Side(style='thin'))

            # Goes over each cell and applies border
            rows = sheet.iter_rows(full_cord)
            for r in rows:
                for cell in r:
                    cell.border = thin_border

        # Creates a dictionary based on a room key
        room_course_dict = set_room_dict(self.list_dict_courses)

        # Goes over room dict
        for key, value in room_course_dict.items():
            if not (key == "ONLINE" or key == "NONE" or key == "ARR"):
                if key in self.cob_rooms:
                    pass
                else:
                    # Marks a room if it is not a part of a college of business rooms.
                    clr = PatternFill(start_color='d5d8e0', end_color='d5d8e0', fill_type='solid')
                    self.sheet["A" + str(start_excel_row)].fill = clr
                    self.sheet["A" + str(start_excel_row)].comment = \
                        Comment("This room is not a part of a College of Business rooms", author="TableMaker")
                # Inserts rooms
                set_room_number(self.sheet, start_excel_row, key)

                days_len = len(self.days)
                for d in range(days_len):
                    # Inserts days
                    set_days_row(self.sheet, start_excel_row, self.days, d)
                    start_excel_row += 1
                for l in range(len(value)):
                    get_day = start_excel_row - days_len
                    # Merging cells with a bold
                    self.merge_excel_cells(get_day, 1, start_excel_row - 1, 1, True, bold=True)
                    while get_day != start_excel_row:
                        if any(c in get_cell_value('B', self.sheet, get_day) for c in (value[l].get('Course_Days'))):
                            for t in range(len(self.list_unique_times)):
                                column = ''.join(string.ascii_uppercase[self.list_unique_times[t].get("Column_Num")])
                                row = str(get_day)
                                if self.list_unique_times[t].get("Time") == value[l].get("Start_Time"):
                                    for en in self.list_unique_times:
                                        if en.get("Time") == value[l].get("End_Time"):
                                            value[l].setdefault("Cell", []).append(column + row + ":" + ''.join(
                                                string.ascii_uppercase[en.get("Column_Num")] + row))

                        get_day += 1

                    # Checking before merging
                    if value[l].get("Cell") is not None:
                        if len(value[l].get("Cell")) < 2:
                            first_cell = re.split('(\d+)', value[l].get("Cell")[0])

                            if (ord(first_cell[0]) - 65) > (ord(first_cell[2][1:]) - 65):
                                first_cell[2] = ":" + first_cell[0]

                                self.sheet.merge_cells(first_cell[0] + first_cell[1] + first_cell[2] + first_cell[3])
                                if self.sheet[first_cell[0] + first_cell[1]].value is None:
                                    if value[l].get("Time_Comment") is None:
                                        self.sheet[first_cell[0] + first_cell[1]] = value[l].get("Course")
                                    else:
                                        self.sheet[first_cell[0] + first_cell[1]] = \
                                            value[l].get("Course") + value[l].get("Time_Comment")
                                else:
                                    if value[l].get("Time_Comment") is None:
                                        self.sheet[first_cell[0] + first_cell[1]] = \
                                            self.sheet[first_cell[0] + first_cell[1]].value + " / " + \
                                            value[l].get("Course")
                                    else:
                                        self.sheet[first_cell[0] + first_cell[1]] = \
                                            self.sheet[first_cell[0] + first_cell[1]].value + " / " + \
                                            value[l].get("Course") + value[l].get("Time_Comment")
                                self.color_cell(value[l].get("Course"), first_cell[0] + first_cell[1])
                            else:
                                self.sheet.merge_cells(first_cell[0] + first_cell[1] + first_cell[2] + first_cell[3])
                                if self.sheet[first_cell[0] + first_cell[1]].value is None:
                                    if value[l].get("Time_Comment") is None:
                                        self.sheet[first_cell[0] + first_cell[1]] = value[l].get("Course")
                                    else:
                                        self.sheet[first_cell[0] + first_cell[1]] = value[l].get("Course") \
                                                                                    + value[l].get("Time_Comment")
                                else:
                                    if value[l].get("Time_Comment") is None:
                                        self.sheet[first_cell[0] + first_cell[1]] = \
                                            self.sheet[first_cell[0] + first_cell[1]].value + " / " + \
                                            value[l].get("Course")
                                    else:
                                        self.sheet[first_cell[0] + first_cell[1]] = \
                                            self.sheet[first_cell[0] + first_cell[1]].value + " / " + \
                                            value[l].get("Course") + "\n" + value[l].get("Time_Comment")
                                self.color_cell(value[l].get("Course"), first_cell[0] + first_cell[1])

                            pass
                        else:
                            # If days happens twice a week
                            if len(value[l].get("Cell")) == 2:
                                first_cell = re.split('(\d+)', value[l].get("Cell")[0])
                                second_cell = re.split('(\d+)', value[l].get("Cell")[1])

                                # Swapping if first cell not alphabetic order
                                if (ord(first_cell[0]) - 65) > (ord(second_cell[2][1:]) - 65):
                                    if (ord(first_cell[0]) - 65) - (ord(second_cell[2][1:]) - 65) > 6:
                                        first_cell[2] = ":" + first_cell[0]
                                        second_cell[0] = first_cell[0]
                                        second_cell[2] = first_cell[2]
                                    else:
                                        first_cell[0] = second_cell[2][1:]
                                        first_cell[2] = second_cell[0]
                                # Merging courses if they have same time and same room
                                if (int(first_cell[1]) == int(second_cell[1])-1) & \
                                        (int(first_cell[3]) == int(second_cell[3])-1):
                                    self.sheet.merge_cells(first_cell[0]+first_cell[1]+second_cell[2]+second_cell[3])
                                    # Checks if cells are empty
                                    if self.sheet[first_cell[0] + first_cell[1]].value is None:
                                        # Checks if Time_Comment existed for this course
                                        if value[l].get("Time_Comment") is None:
                                            self.sheet[first_cell[0] + first_cell[1]] = value[l].get("Course")
                                        else:
                                            self.sheet[first_cell[0] + first_cell[1]] = \
                                                value[l].get("Course") + "\n" + value[l].get("Time_Comment")
                                    else:
                                        # Checks if Time_Comment existed for this course
                                        if value[l].get("Time_Comment") is None:
                                            self.sheet[first_cell[0] + first_cell[1]] = \
                                                self.sheet[first_cell[0] + first_cell[1]].value + " / " + \
                                                value[l].get("Course")
                                        else:
                                            self.sheet[first_cell[0] + first_cell[1]] = \
                                                self.sheet[first_cell[0] + first_cell[1]].value + " / " + \
                                                value[l].get("Course") + "\n" + value[l].get("Time_Comment")
                                    # Colors the cell by specific color
                                    self.color_cell(value[l].get("Course"), first_cell[0]+first_cell[1])
                                else:
                                    # Merging cells if they don't have same times and same room
                                    self.sheet.merge_cells(first_cell[0] + first_cell[1]
                                                           + first_cell[2] + first_cell[3])
                                    self.sheet.merge_cells(second_cell[0] + second_cell[1] +
                                                           second_cell[2] + second_cell[3])
                                    if value[l].get("Time_Comment") is None:
                                        self.sheet[first_cell[0] + first_cell[1]] = value[l].get("Course")
                                        self.sheet[second_cell[0] + second_cell[1]] = value[l].get("Course")
                                        self.color_cell(value[l].get("Course"), first_cell[0] + first_cell[1])
                                        self.color_cell(value[l].get("Course"), second_cell[0] + second_cell[1])
                                    else:
                                        self.sheet[first_cell[0] + first_cell[1]] = value[l].get("Course") + "\n" + \
                                                                                    value[l].get("Time_Comment")
                                        self.sheet[second_cell[0] + second_cell[1]] = \
                                            value[l].get("Course") + "\n" + value[l].get("Time_Comment")
                                        self.color_cell(value[l].get("Course"), first_cell[0] + first_cell[1])
                                        self.color_cell(value[l].get("Course"), second_cell[0] + second_cell[1])

                            elif len(value[l].get("Cell")) == 3:
                                # In progress
                                # first_cell = re.split('(\d+)', value[l].get("Cell")[0])
                                # second_cell = re.split('(\d+)', value[l].get("Cell")[1])
                                # third_cell = re.split('(\d+)', value[l].get("Cell")[2])
                                pass

        border_all_cells(self.sheet)

    def color_cell(self, text, coordinate, course_type_list=True):
        """Colors a course based on a department color"""
        # Different colors for each department
        # Accounting
        acct = PatternFill(start_color='FF958C', end_color='FF958C',
                           fill_type='solid')
        # Business Law
        blaw = PatternFill(start_color='FFCC00', end_color='FFCC00',
                           fill_type='solid')
        # Business
        bus = PatternFill(start_color='FFFF00', end_color='FFFF00',
                          fill_type='solid')
        # Finance
        fin = PatternFill(start_color='99CC00', end_color='99CC00',
                          fill_type='solid')
        # International Business
        ibus = PatternFill(start_color='8CF6FF', end_color='8CF6FF',
                           fill_type='solid')
        # Master of Business Administration
        mba = PatternFill(start_color='33CCCC', end_color='33CCCC',
                          fill_type='solid')
        # Master of Accounting
        macc = PatternFill(start_color='FF00FF', end_color='FF00FF',
                           fill_type='solid')
        # Management
        mgmt = PatternFill(start_color='CC99FF', end_color='CC99FF',
                           fill_type='solid')
        # Marketing
        mrkt = PatternFill(start_color='A28CFF', end_color='A28CFF',
                           fill_type='solid')

        # Checks for the first two letters to identify the color
        # Accounting
        if text[:2] == "AC":
            color = acct
            if course_type_list is True:
                self.course_types_list.append("ACCT")
        # Business Law
        elif text[:2] == "BL":
            color = blaw
            if course_type_list is True:
                self.course_types_list.append("BLAW")
        # Business
        elif text[:2] == "BU":
            color = bus
            if course_type_list is True:
                self.course_types_list.append("BUS")
        # Finance
        elif text[:2] == "FI":
            color = fin
            if course_type_list is True:
                self.course_types_list.append("FIN")
        # International Business
        elif text[:2] == "IB":
            color = ibus
            if course_type_list is True:
                self.course_types_list.append("IBUS")
        # Master of Business Administration
        elif text[:2] == "MB":
            color = mba
            if course_type_list is True:
                self.course_types_list.append("MBA")
        # Master of Accounting
        elif text[:2] == "MA":
            color = macc
            if course_type_list is True:
                self.course_types_list.append("MACC")
        # Management
        elif text[:2] == "MG":
            color = mgmt
            if course_type_list is True:
                self.course_types_list.append("MGMT")
        # Marketing
        elif text[:2] == "MR":
            color = mrkt
            if course_type_list is True:
                self.course_types_list.append("MRKT")
        else:
            color = PatternFill(start_color='eeefef', end_color='eeefef',
                                fill_type='solid')
        # Fills the color
        self.sheet[coordinate].fill = color
        # Makes the text be in the center of a cell
        self.sheet[coordinate].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    def color_cell_meaning(self):
        """Will appear on a right side which color assign to which department"""
        get_max_column = self.sheet.max_column
        row = 2

        def remove_duplicates(course_list):
            return list(set(course_list))

        def sort_list(course_list):
            course_list.sort()
            return course_list

        unique_types = sort_list(remove_duplicates(self.course_types_list))
        #print(unique_types)

        for i in range(len(unique_types)):
            alphabet = ''.join(string.ascii_uppercase[get_max_column+1])
            self.color_cell(unique_types[i], alphabet+str(row), False)
            self.sheet[''.join(string.ascii_uppercase[get_max_column+2])+str(row)] = "-" + unique_types[i]
            row += 1

    def set_page_break(self):
        # 40 rows per page
        get_max_row = self.sheet.max_row
        get_max_column = self.sheet.max_column

        if len(self.days) == 4:
            if get_max_row >= 40:
                while get_max_row >= 40:
                    self.sheet.sheet_properties.pageSetUpPr.fitToPage = True
                    openpyxl.worksheet.pagebreak.PageBreak.tagname = 'rowBreaks'
                    page_break_row = Break((get_max_row + 1)-37)
                    self.sheet.page_breaks.append(page_break_row)

                    openpyxl.worksheet.pagebreak.PageBreak.tagname = 'colBreaks'
                    page_break_column = Break(get_max_column + 1)
                    self.sheet.page_breaks.append(page_break_column)
                    get_max_row -= 37

            elif get_max_row == 40:
                pass
            else:
                self.sheet.sheet_properties.pageSetUpPr.fitToPage = True
                openpyxl.worksheet.pagebreak.PageBreak.tagname = 'rowBreaks'
                page_break_row = Break(get_max_row + 1)
                self.sheet.page_breaks.append(page_break_row)

                openpyxl.worksheet.pagebreak.PageBreak.tagname = 'colBreaks'
                page_break_column = Break(get_max_column + 1)
                self.sheet.page_breaks.append(page_break_column)









