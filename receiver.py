import datetime
import os
import re
import sys
import threading
import time
from itertools import chain
from tkinter import messagebox

import openpyxl
from openpyxl.comments import Comment
from openpyxl.styles import PatternFill

import tableDesign


class DataProcessor:

    def __init__(self, file_directory, table_name, table_semester, table_year, table_type, friday):
        # Getting user information
        self.file_directory = file_directory
        self.table_name = table_name
        self.table_semester = table_semester
        self.table_year = table_year
        self.days_order = table_type
        self.friday_choice = friday

        self.excel_data_list = None
        self.user_excel_errors = []
        self.list_file_paths = []
        self.list_dict_courses = []

        # Temporary Values. Will change dates based on a user request in the future
        self.days = ["Monday", "Wednesday", "Tuesday", "Thursday"]

        # User-Excel-Columns identification
        self.excel_course_name = 1
        self.excel_course_number = 2
        self.excel_course_section = 3
        self.excel_course_credits = 4
        self.excel_course_title = 5
        self.excel_course_room = 6
        self.excel_course_days = 7
        self.excel_course_time = 8

        # Start-user-data
        self.start_row = 4

        # In case of any errors, it will give three chances before the program stops
        self.number_close_trials = 0

        self.main_program_order()

    def main_program_order(self):
        """Main program logic"""

        def create_excel_copies():
            """Creates a folder to store all the copy files"""
            if not os.path.exists('copy_folder'):
                # Creates a folder for files
                os.makedirs('copy_folder')

        def close_file_error():
            user_response = messagebox.showerror("Close File", "Save and close excel files to continue... \n\n"
                                                 + str(p))
            return user_response

        try:
            create_excel_copies()
            for i in range(len(self.file_directory)):
                get_file_name = os.path.basename(self.file_directory[i])
                workbook_copy = openpyxl.load_workbook(self.file_directory[i])
                set_file_path = 'copy_folder\\'+'copy_' + get_file_name
                self.get_excel_data(workbook_copy, self.file_directory[i])
                self.set_dict_courses(self.excel_data_list)
                self.time_conflict()
                self.list_file_paths.append(set_file_path)
                self.color_comment_copy_excel(workbook_copy, self.file_directory[i])
                workbook_copy.save(set_file_path)
            self.get_excel_errors()
            self.create_excel_table()
        except PermissionError as p:
            # Gives a user three chances to close excel files
            if self.number_close_trials > 2:
                self.user_excel_errors = 'User_Doesnt_Listen'
                t = threading.Thread(target=sys.exit)
                t.setDaemon(True)
                t.start()
            else:
                msg_response = close_file_error()
                if msg_response == "ok":
                    self.number_close_trials += 1
                    self.main_program_order()

    def get_excel_data(self, wb_copy, file_path):
        """Gets all information from excel file"""
        excel_data = []
        for i in range(len(wb_copy.worksheets)):
            get_excel_workbook = openpyxl.load_workbook(file_path, read_only=True)

            get_sheet_name = get_excel_workbook.get_sheet_names()[i]
            read_mode_worksheet = get_excel_workbook[get_sheet_name]
            user_data = list(self.iter_rows(read_mode_worksheet, file_path, get_sheet_name))
            excel_data.append(user_data)
        self.excel_data_list = (list(chain.from_iterable(excel_data)))

    def create_report_dictionary(self, excel_row, excel_column, excel_file_path, excel_sheet_name, error_color,
                                 error_comment=None):
        """If there is any conflict, error, missing data it will store that information in user_excel_errors
        dictionary"""

        report_dict = dict()
        report_dict["Row"] = excel_row
        report_dict["Column"] = excel_column
        report_dict["File_Path"] = excel_file_path
        report_dict["Sheet_Name"] = excel_sheet_name
        report_dict["Color"] = error_color

        if error_comment is None:
            pass
        else:
            report_dict["Comment"] = error_comment

        self.user_excel_errors.append(report_dict.copy())

    def iter_rows(self, read_worksheet, file_path, sheet_name):
        """Loops through rows. Checks if the value is none. And transforms time to the same format"""

        def convert_time(user_time):
            """Converts time into the correct format"""
            try:
                user_time = datetime.datetime.strptime(user_time, "%H:%M")
                converted_time = datetime.datetime.strftime(user_time, '%I:%M')
                return converted_time
            except ValueError:
                try:
                    user_time = datetime.datetime.strptime(user_time, "%H")
                    converted_time = datetime.datetime.strftime(user_time, '%I:%M')
                    return converted_time
                except ValueError:
                    try:
                        user_time = datetime.datetime.strptime(user_time, "%H:%M %p")
                        converted_time = datetime.datetime.strftime(user_time, '%I:%M')
                        return converted_time
                    except ValueError:
                        try:
                            user_time = datetime.datetime.strptime(user_time, "%H:%M%p")
                            converted_time = datetime.datetime.strftime(user_time, '%I:%M')
                            return converted_time
                        except ValueError:
                            pass

        data_start_row = self.start_row

        iterable_row = read_worksheet.iter_rows()

        # Skips unnecessary rows
        for i in range(data_start_row-1):
            next(iterable_row)

        for row in iterable_row:
            # Creating temporary list
            excel_cell_value = []
            for cell in row:
                try:
                    if cell.value is None:
                        excel_cell_value.append(cell.value)
                    elif (cell.column == self.excel_course_time) & (cell.value in {"Online", "ONLINE", "online"}):
                        excel_cell_value.append("Online-Online")
                    elif cell.column == self.excel_course_time:
                        try:

                            split_time = cell.value.split('-')

                            split_time[0] = "".join(split_time[0].split())
                            split_time[1] = "".join(split_time[1].split())

                            # Converts our time to the same format
                            converted_time_first = convert_time(split_time[0])
                            converted_time_second = convert_time(split_time[1])
                            excel_cell_value.append(converted_time_first + "-" + converted_time_second)
                        except AttributeError:
                            # Will mark if program can't read the datetime
                            if type(cell.value) is datetime.time:
                                excel_cell_value.append('%s:%s-None' % (cell.value.hour, cell.value.minute))

                                comment = "Does it contain the start time or end time of the course?"
                                self.create_report_dictionary(cell.row, cell.column, file_path, sheet_name,
                                                              error_color="FF687B", error_comment=comment)
                            elif cell.value is not None:

                                comment = "Does it contain the start time or end time of the course?"
                                self.create_report_dictionary(cell.row, cell.column, file_path, sheet_name,
                                                              error_color="FF687B", error_comment=comment)

                            else:
                                excel_cell_value.append("None-None")
                        except IndexError:
                            pass
                    else:
                        excel_cell_value.append(cell.value)
                except AttributeError:
                    pass

            if all(v is None for v in excel_cell_value):
                data_start_row += 1
            else:
                excel_cell_value.insert(0, file_path)
                excel_cell_value.insert(0, sheet_name)
                excel_cell_value.insert(0, data_start_row)

                data_start_row += 1
            yield excel_cell_value

    def color_comment_copy_excel(self, workbook, file_path):
        """Will comment and color excel copy sheets where the program finds any mistakes or conflicts.
                It will help the user to find them faster than ever."""

        get_excel_workbook = workbook
        # Goes through each error
        for i in range(len(self.user_excel_errors)):
            # Checks if this error is assign to this file
            if self.user_excel_errors[i].get("File_Path") == file_path:
                edit_sheet = get_excel_workbook[self.user_excel_errors[i].get("Sheet_Name")]
                # Takes error information
                row_num = self.user_excel_errors[i].get("Row")
                column_num = self.user_excel_errors[i].get("Column")
                error_color = self.user_excel_errors[i].get("Color")
                # Fills cell with error color
                edit_sheet.cell(row=row_num, column=column_num).fill = PatternFill(start_color=error_color,
                                                                                   end_color=error_color,
                                                                                   fill_type='solid')
                # Comments cell
                if self.user_excel_errors[i].get("Comment") is not None:
                    comment = self.user_excel_errors[i].get("Comment")
                    edit_sheet.cell(row=row_num, column=column_num).comment = Comment(comment, author="TableMaker")
                else:
                    pass

    def mark_none_values(self, data_list):
        """Marks none values with specific color"""
        for i in range(len(data_list)):
            # Goes every None value
            none_index = [i for i, e in enumerate(data_list[i]) if e == "None"]
            for val_length in range(len(none_index)):
                column = none_index[val_length] - 2
                # if column is above 13 we don't need to mark
                if column > 13:
                    pass
                else:
                    try:
                        self.create_report_dictionary(data_list[i][0], column, data_list[i][2],
                                                      data_list[i][1], error_color="FADDA7")
                    except IndexError:
                        pass

    def set_dict_courses(self, excel_data):
        """Inserting all information into a dictionary"""

        def clear_unnecessary_list(data_list):
            data_list = list(filter(None, data_list))  # Deletes empty lists
            # Deletes None lists
            data_list = [None if list(set(v)) == [None] else v for v in data_list]
            data_list = [v for v in data_list if v is not None]
            data_list = [['None' if v is None else v for v in row] for row in data_list]
            return data_list

        def course_room_format(room_number):
            """Formats room to follow the same format"""
            str_list = re.split('(\d+)', room_number)
            # Removes empty str
            filter_str_list = list(filter(None, str_list))
            for f in range(len(filter_str_list)):
                filter_str_list[f] = filter_str_list[f].lower()
                filter_str_list[f] = ''.join(filter_str_list[f].split())
                filter_str_list[f] = filter_str_list[f].upper()
                if f == 1:
                    while len(filter_str_list[f]) < 4:
                        filter_str_list[f] = '0' + filter_str_list[f]
            return ' '.join(filter_str_list)

        def convert_user_days(day, k):
            """Transfers day into proper format"""
            # Transforms to Monday
            if any(c in day[k].upper() for c in "M"):
                return "Monday"
            elif any(c in day[k:3].upper() for c in "MON"):
                return "Monday"
            elif any(c in day[k:6].upper() for c in "MONDAY"):
                return "Monday"

            # Transforms to Thursday
            elif any(c in day[k].upper() for c in ("R", "H")):
                return "Thursday"
            elif day[k:2].upper() == "TH":
                return "Thursday"
            elif day[k:3].upper() == "THU":
                return "Thursday"
            elif day[k:7].upper() == "THURSDAY":
                return "Thursday"

            # Transforms to Tuesday
            elif any(c in day[k].upper() for c in "T"):
                return "Tuesday"
            elif any(c in day[k:3].upper() for c in "TUE"):
                return "Tuesday"
            elif any(c in day[k:7].upper() for c in "TUESDAY"):
                return "Tuesday"

            # Transforms to Wednesday
            elif any(c in day[k].upper() for c in "W"):
                return "Wednesday"
            elif any(c in day[k:3].upper() for c in "WED"):
                return "Wednesday"
            elif any(c in day[k:7].upper() for c in "WEDNESDAY"):
                return "Wednesday"

            # Transforms to Friday
            elif any(c in day[k].upper() for c in "F"):
                return "Friday"
            elif any(c in day[k:3].upper() for c in "FRI"):
                return "Friday"
            elif any(c in day[k:7].upper() for c in "FRIDAY"):
                return "Friday"

            else:
                pass

        excel_data = clear_unnecessary_list(excel_data)
        self.mark_none_values(excel_data)

        def set_online_course(d_courses, l_data):
            """Sets online courses to a specific type."""
            d_courses["Course_Days"] = []
            d_courses["Start_Time"] = "Online"
            d_courses["End_Time"] = "Online"
            d_courses["Type"] = "Online"
            d_courses["Row"] = l_data[0]
            d_courses["File"] = l_data[2]
            d_courses["Sheet_Name"] = l_data[1]

            return d_courses

        for j in excel_data:
            try:
                dict_courses = dict()
                time_split = j[self.excel_course_time + 2].split('-')

                # Sets course to specific format.
                dict_courses["Course"] = str(j[self.excel_course_name + 2]).upper() + " " + str(
                    j[self.excel_course_number + 2]) + "-" + str(j[self.excel_course_section + 2])
                # Sets room to specific format.
                dict_courses["Room"] = course_room_format(str(j[self.excel_course_room + 2]))

                # Creates a dictionary
                if dict_courses.get("Room") == "ONLINE":
                    dict_courses = set_online_course(dict_courses, j)
                elif time_split[0].upper() == "ONLINE":
                    dict_courses["Room"] = course_room_format(str(j[self.excel_course_room + 2]))
                    dict_courses = set_online_course(dict_courses, j)
                else:
                    dict_courses["Room"] = course_room_format(str(j[self.excel_course_room + 2]))
                    dict_courses["Course_Days"] = []
                    dict_courses["Row"] = j[0]
                    dict_courses["File"] = j[2]
                    dict_courses["Sheet_Name"] = j[1]
                    dict_courses["Type"] = "Classroom"
                    dict_courses["Start_Date"] = j[13]
                    dict_courses["End_Date"] = j[14]
                    try:
                        dict_courses["Start_Time"] = time_split[0]
                        dict_courses["End_Time"] = time_split[1]

                        # Due to space table concern, we need to limit evening classes cells
                        if time_split[0][1] == '6':
                                dict_courses["Time_Comment"] = " ends at " + time_split[0]
                                dict_courses["End_Time"] = time_split[0]
                        if time_split[1][1] == '6' or time_split[1][1] == '7':
                            dict_courses["Time_Comment"] = " ends at " + time_split[1]
                            dict_courses["End_Time"] = "06:00"

                    except IndexError:
                        # Marks a course if program couldn't read it
                        dict_courses["Start_Time"] = "None"
                        dict_courses["End_Time"] = "None"
                        dict_courses["Type"] = "Error"
                    # Splits days and converts to the proper format
                    if "," in j[self.excel_course_days + 2]:
                        split_by_comma = [x.strip() for x in j[self.excel_course_days + 2].split(',')]
                        for l in range(len(split_by_comma)):
                            dict_courses["Course_Days"].append(convert_user_days(split_by_comma[l], 0))

                    elif "None" in j[self.excel_course_days + 2]:
                        pass

                    else:
                        for i in range(len(j[self.excel_course_days + 2])):
                            dict_courses["Course_Days"].append(convert_user_days(j[self.excel_course_days + 2], i))

                self.list_dict_courses.append(dict_courses.copy())
            except AttributeError:
                # If an error occurred, it will mark the whole row
                for i in range(12):
                    self.create_report_dictionary(j[0], i+1, j[2], j[1], error_color="FF687B")
                # Marks last cell with comment
                comment = "A program couldn't read this row correctly. Report it if needed."
                self.create_report_dictionary(j[0], 13, j[2], j[1], error_color="FF687B", error_comment=comment)

    def time_conflict(self):
        """Loops through each dictionary in the list. Looks for similar rooms and days.
        Finds a time conflict between courses. Deletes the conflict dict."""

        # Creates a copy of our main dict
        list_dict = self.list_dict_courses.copy()
        for ig in range(len(list_dict)):
            for d in range(len(list_dict) - 1):
                if ig != (d + 1):
                    try:
                        # Checks Course days similarities
                        if not list(set(list_dict[ig].get("Course_Days")) & set(list_dict[d + 1].get("Course_Days"))):
                            pass
                        else:
                            # Checks if online
                            if list_dict[ig].get('Room') == 'Online' or list_dict[ig].get("Start_Time") == 'Online':
                                pass
                            elif list_dict[d + 1].get('Room') == 'Online' or list_dict[d + 1].get(
                                    "Start_Time") == 'Online':
                                pass

                            else:
                                start_time_i = list_dict[ig].get('Start_Time')
                                start_time_d = list_dict[d + 1].get('Start_Time')
                                end_time_i = list_dict[ig].get('End_Time')
                                end_time_d = list_dict[d + 1].get('End_Time')

                                if start_time_i is None or start_time_d is None:
                                    pass
                                elif start_time_i == 'None' or start_time_d == 'None':
                                    pass
                                elif end_time_i is None or end_time_d is None:
                                    pass
                                elif end_time_i == 'None' or end_time_d == 'None':
                                    pass
                                else:
                                    if (list_dict[ig].get('Room') is not None) & \
                                            (list_dict[d + 1].get('Room') is not None):
                                        room_ig = "".join(list_dict[ig].get('Room').split())
                                        room_d = "".join(list_dict[d + 1].get('Room').split())

                                        section_numb_ig = list_dict[ig].get('Course').split("-")
                                        section_numb_d = list_dict[d + 1].get("Course").split("-")

                                        # It is normal for section 40 or 41 to conflict with another course
                                        if (section_numb_d[1] == '40') or (section_numb_ig[1] == '40'):
                                            pass
                                        elif (section_numb_d[1] == '41') or (section_numb_ig[1] == '41'):
                                            pass

                                        elif room_ig == room_d:

                                            # Transforms variables to float
                                            start_time_i = (float(start_time_i[0:2] + '.' + start_time_i[3:5]))
                                            start_time_d = (float(start_time_d[0:2] + '.' + start_time_d[3:5]))
                                            end_time_i = (float(end_time_i[0:2] + '.' + end_time_i[3:5]))
                                            end_time_d = (float(end_time_d[0:2] + '.' + end_time_d[3:5]))

                                            # Checks for conflicts
                                            if start_time_d <= start_time_i <= end_time_d:
                                                self.time_conflict_comment(list_dict[ig], list_dict[d + 1])
                                                del self.list_dict_courses[ig]

                                            elif start_time_d <= end_time_i <= end_time_d:
                                                self.time_conflict_comment(list_dict[ig],
                                                                           list_dict[d + 1])
                                                del self.list_dict_courses[ig]
                                            else:
                                                pass
                                                # Currently not working properly. So it will get fixed later.
                                                """
                                                fifteenth_minutes_start_i = start_time_i + 0.14
                                                fifteenth_minutes_end_i = end_time_i + 0.14
                                                fifteenth_minutes_start_d = start_time_d + 0.14
                                                fifteenth_minutes_end_d = end_time_d + 0.14

                                                def check_minutes(time):
                                                    if math.modf(time)[0] >= 0.60:
                                                        x = math.modf(time)[1] + 1 + math.modf(time)[0] - 0.60
                                                        return x
                                                    else:
                                                        return time

                                                fifteenth_minutes_start_i = check_minutes(fifteenth_minutes_start_i)
                                                fifteenth_minutes_end_i = check_minutes(fifteenth_minutes_end_i)
                                                fifteenth_minutes_start_d = check_minutes(fifteenth_minutes_start_d)
                                                fifteenth_minutes_end_d = check_minutes(fifteenth_minutes_end_d)
                                                print("Nnn")
                                                print(fifteenth_minutes_start_i)
                                                print(fifteenth_minutes_end_i)
                                                print(fifteenth_minutes_start_d)
                                                print(fifteenth_minutes_end_d)
                                                if fifteenth_minutes_start_d <= fifteenth_minutes_start_i <= 
                                                fifteenth_minutes_end_d:
                                                    self.time_conflict_comment(list_dict[ig], list_dict[d + 1], True)
                                
                                                else:
                                                    if fifteenth_minutes_start_d <= fifteenth_minutes_end_d <= 
                                                    fifteenth_minutes_end_i:
                                                        self.time_conflict_comment(list_dict[ig],
                                                                                   list_dict[d + 1], True)
                                                """

                    except IndexError:
                        pass

    def time_conflict_comment(self, first_dict, second_dict, fifteenth_conflict=False):
        """Creates a dictionary if program finds a conflict"""
        if fifteenth_conflict is True:
            comment = second_dict.get("Course") + ": " + "time difference is less than 15 min." + (" " * 164)
        else:
            comment = second_dict.get("Course") + ": " + "conflicts with another course." + (" " * 173)

        self.create_report_dictionary(
            excel_row=first_dict.get("Row"),
            excel_column=1,
            excel_file_path=first_dict.get("File"),
            excel_sheet_name=first_dict.get("Sheet_Name"),
            error_color="FF687B", error_comment=comment)

        self.create_report_dictionary(
            excel_row=first_dict.get("Row"),
            excel_column=5,
            excel_file_path=first_dict.get("File"),
            excel_sheet_name=first_dict.get("Sheet_Name"),
            error_color="FF687B", error_comment=comment)

    def create_excel_table(self):
        """Moves into another class"""
        tableDesign.MasterDesign(self.list_dict_courses, self.days, self.table_year, self.table_name,
                                 self.table_semester)

    def get_excel_errors(self):
        """Returns founded errors"""
        return self.user_excel_errors


start_time = time.time()
print("--- %s seconds ---" % (time.time() - start_time))

